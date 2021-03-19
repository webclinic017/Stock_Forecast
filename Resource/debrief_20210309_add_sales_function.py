# -*- coding: utf-8 -*-
"""
Created on Mon Jan 25 17:24:39 2021

@author: 0108575
"""

#%%# 輸入區

url = 'https://docs.google.com/spreadsheets/d/1crmJPaznfu0tqHD97eq1yHC22mhuWe0jpItbnYEWlf4/edit?usp=sharing'
chart_name = 'Debrief'
data_path = 'D:/tool/'
path = r'D:\Data_Mining\Projects\DM-Debrif'
author = 'Aron Wu #739'


# 要計算的分店
exclude_stores=[4,5,8,19]
include_stores=[1]







# Worklist
# 讀取會員狀態資料，目前先用member_data
# status_flie_name = '20201209.rds'



#%%# 引入區(有需要再動)
import sys,pandas as pd, arrow, numpy as np, os, smtplib
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
sys.path.append(r'D:\Data_Mining\Projects\Codebase_YZ')

from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart


# Codebase
path_codebase = ['D:/Data_Mining/Projects/Codebase_YZ']

for i in path_codebase:    
    if i not in sys.path:
        sys.path = [i] + sys.path

    
import codebase_yz as cbyz
import codebase_dm as cbdm
import toolbox as t



path_resource = path + '/Resource'
path_function = path + '/Function'
path_temp = path + '/Temp'
path_export = path + '/Export'


cbyz.os_create_folder(path=[path_resource, path_function, 
                            path_temp, path_export])      








def df_condition(df_name, label, condition, takeout_value=''):
    '''
    dataframe的快速條件選擇器。
    '''
    if takeout_value == '':
        text = ("{}[{}['{}']{}]").format(df_name, df_name, label, condition)
        result = eval(text) 
    else:    
        text = ("{}[{}['{}']{}]['{}']") \
                .format(df_name, df_name, label, condition, takeout_value)
        result = list(eval(text))[0]
    
    return result

def takeout_first(Series):
    return list(Series)[0]

def now():
    return arrow.now().format('YYYY-MM-DD HH:mm:ss ')

def get_dm_info(prom_type=['nDM', 'nFlyer', 'HB'],
                begin_date='20170101',
                end_date=arrow.now().shift(months=2).format('YYYYMMDD'),
                add_sub_channel = False):
    '''
    取得DM的資訊。
    prom_type       :list 選擇促銷DM類型，預設為['nDM', 'nFlyer', 'HB']
    begin_date      :str  要抓取的開始日期，格式為YYYYMMDD，預設為20170101
    end_date        :str  要抓取的結束時間，格式為YYYYMMDD，預設為使用當日往後抓2個月
    add_sub_channel :bool 是否要顯示sub_channel資訊，預設為False
    return          :dataframe DM的資訊表。
    '''
    
    prom_type_already = str()
    if type(prom_type) == list:
        if len(prom_type) == 0 :
            prom_type_already=t.list_to_string(['nDM', 'nFlyer', 'HB'],True)
        else:
            prom_type_already=t.list_to_string(prom_type,True)
    else:
        prom_type_already="'"+prom_type+"'"
           
    #抓取資料
    sql=('''
    select prom_no, prom_type, prom_name, 
    to_char(begin_date, 'yyyymmdd') begin_date, 
    to_char(end_date, 'yyyymmdd') end_date 
    from rdba.prom_slip_dm 
    where prom_type in ({})
    and ((to_char(begin_date, 'yyyymmdd') <= {} and to_char(end_date, 'yyyymmdd') >= {} ) or
         (to_char(begin_date, 'yyyymmdd') >= {} and to_char(end_date, 'yyyymmdd') <= {} ) or
         (to_char(begin_date, 'yyyymmdd') <= {} and to_char(end_date, 'yyyymmdd') >= {} ) or
         (to_char(begin_date, 'yyyymmdd') <= {} and to_char(end_date, 'yyyymmdd') >= {}))
    order by prom_no desc
    ''').format(prom_type_already, \
        begin_date,end_date, \
        begin_date,end_date, \
        begin_date,begin_date, \
        end_date,end_date).replace('\n',' ')
        
    data = t.query_sql(sql)
    
    data["CHANNEL_ID"]=1

    #處理是否要增加sub_channel的欄位
    if add_sub_channel == True :
        data["SUB_CHANNEL"]=np.select(
            [data["PROM_TYPE"]=='nDM',
             data["PROM_TYPE"]=='nFlyer',
             data["PROM_TYPE"]=='HB'],
            [1,2,3],default=0)
        
    data = data.sort_values(by='END_DATE',ascending=False)
    
    return data

store_list = t.query_sql('select STORE_NO from store@csm where store_no not in (4,5,8,19,0)')['STORE_NO'].to_list()


#%% 處理全商品表
item_data = t.query_sql('''
                        select division_cd division_id,section_cd section_id,
                        item_class_cd group_id,
                        item_id item_no,item_name
                        from pmart.gnl_dim_item@dw
                        ''')

division = t.query_sql('''
                       select division_no division_id , name division_name from divisions@csm
                       ''')

section = t.query_sql('''
                       select SECTION_NO SECTION_ID,NAME SECTION_NAME,EN_NAME SECTION_EN_NAME from sections@csm
                       ''')               

series = t.query_sql('''
                      select tb1.series_id, series_name, item_id item_no
                      from (select series_id, item_id 
                            from itm_series_detail@rtmm) tb1,  
                          (select series_id, series_name
                           from itm_series_header@rtmm ) tb2
                      where tb1.series_id = tb2.series_id (+)''')

cluster = t.query_sql('''
                       select tb1.cluster_id, tb1.cluster_name,tb2.item_no
                       from (select cluster_id, name cluster_name from itm_cluster@rtmm ) tb1,
                       (select cluster_id, item_id item_no   from itm_cluster_item@rtmm ) tb2
                       where tb1.cluster_id (+) = tb2.cluster_id ''')

all_item_data = item_data.merge( division , how = 'left')\
                         .merge( section , how = 'left')\
                         .merge( series, how = 'left')\
                         .merge( cluster , how = 'left')

all_item_data = all_item_data[all_item_data['DIVISION_ID']!=9]





# %% 主程式
already_check = t.check_data(url)

for i in range(len(already_check)):
    
    todo_data = already_check[i]
    prom_no = todo_data[2]
    start_time = t.spend_start_time()

    timestamp, prom_type, prom_no, sub_channel, \
        last_prom_no, user_mail, channel = todo_data
    
    # Check，檢查google表單 ......
    # channel = channel.split(',')[0]
    if channel == '1.平面媒體':
        channel = 1
    
    # sub_channel = subchannel.split(',')[0]
    if sub_channel == '3.型錄':
        sub_channel = 3      
    
    print(now()+' start '+str(prom_no)+' debrief')
    
    
    
    store_list_raw = cbyz.get_store_list(exclude_stores=exclude_stores,
                                         include_stores=include_stores)

    
    store_list = store_list_raw['STORE_IN_SERIES'].tolist()
    
    # ......
    file_path = path + '/' + str(prom_no) + '/'

    try:
        os.mkdir(file_path)
    except:
        print('資料夾已建立。')
        
    dm_info = get_dm_info(add_sub_channel = True)
    
    
    # 取得檔期資訊
    this_dm_prom_info = df_condition('dm_info','PROM_NO','==str(prom_no)')
    this_dm_prom_info = df_condition('this_dm_prom_info','CHANNEL_ID',
                                     '==int(channel)')
    this_dm_prom_info = df_condition('this_dm_prom_info','SUB_CHANNEL',
                                     '==int(sub_channel)')
    
    this_dm_prom_info = this_dm_prom_info \
                        .reset_index(drop=True)
    
    last_dm_prom_info = df_condition('dm_info', 'PROM_NO',
                                     '==str(last_prom_no)')
    last_dm_prom_info = df_condition('last_dm_prom_info', 'CHANNEL_ID',
                                     '==int(channel)')
    last_dm_prom_info = df_condition('last_dm_prom_info','SUB_CHANNEL',
                                     '==int(sub_channel)')
    
    last_dm_prom_info = last_dm_prom_info \
                        .reset_index(drop=True)    
                        
                        
    # 確認檔期時間
    begin_date = this_dm_prom_info.loc[0, 'BEGIN_DATE']
    begin_date = int(begin_date)
    end_date = this_dm_prom_info.loc[0, 'END_DATE']
    end_date = int(end_date)
    
    
    prom_days = cbyz.ymd(end_date) - cbyz.ymd(begin_date)
    elas_begin_date = cbyz.date_cal(begin_date, amount=-prom_days.days-1, 
                                    unit='d')
    
    elas_end_date = cbyz.date_cal(begin_date, amount=-1, unit='d')    
    

    last_begin_date = last_dm_prom_info.loc[0, 'BEGIN_DATE']
    last_begin_date = int(last_begin_date)
    last_end_date = last_dm_prom_info.loc[0, 'END_DATE']                        
    last_end_date = int(last_end_date)
                        

    #%%#取得檔期商品資料
    print(now()+' '+prom_no+' item start')
    #本檔期商品===============================================================================
    sql = ( """
            select distinct promotion_id prom_no, 
            prom_type item_prom_type, channel_id, sub_channel_id,
            item_id item_no, rule_type , action_id
            from pm_st_item@rtmm
            where promotion_id = {}
            and prom_type in ('G') 
            and channel_id in ({})
            and sub_channel_id in ({})
            and store_id in ({}) 
            """).format(prom_no,
                        channel,
                        sub_channel,
                        t.list_to_string(store_list)).replace('\n',' ')
        
    this_prom_item = t.query_sql(sql)
    
    #處理本檔期CUT_ID提品編號    
    sql = ('''
           select distinct promotion_id prom_no,
           prom_type item_prom_type, channel_id, sub_channel_id,
           item_id item_no, rule_type, action_id, cut_id
           from pm_st_item_plan@rtmm
           where promotion_id = {}
           and prom_type = 'G'
           and channel_id in({})
           and sub_channel_id in ({})
           and store_id in ({})
                ''').format(prom_no,
                            channel,
                            sub_channel,
                            t.list_to_string(store_list)).replace('\n',' ') 
    
    this_prom_cut = t.query_sql(sql)
    this_prom_cut['CUT_ID'] = this_prom_cut['CUT_ID'].fillna(0).apply(float).apply(int) #20210203 會議改為0
    
    #處理本檔期PAGE頁面資訊
    sql = '''
          select promotion_id prom_no, action_id,
          rule_type, sub_dm_id, page
          from pm_st_item_plan_extra@rtmm
          where promotion_id = {}
          '''.format(prom_no)
    
    this_extra_info = t.query_sql(sql)
    this_extra_info['PAGE'] = this_extra_info['PAGE'].fillna(0).apply(int) #20210203 會議改為0
            
    #處理dm_name_data資訊
    sql = '''
             select promotion_id prom_no,sub_dm_id, sub_dm_name
             from pm_sub_dm_tl@rtmm
             where promotion_id = {}
          '''.format(prom_no)
            
    this_dm_name = t.query_sql(sql)
            
    #整合本檔期促銷商品資訊
    this_prom_data = this_prom_item.merge(this_prom_cut,
                                           how ='left' ,
                                           on = ['PROM_NO', 'ITEM_PROM_TYPE',
                                                 'ACTION_ID','CHANNEL_ID', 'SUB_CHANNEL_ID',
                                                 'ITEM_NO','RULE_TYPE'])\
                                    .merge(this_extra_info,
                                           how ='left' , 
                                           on = ['PROM_NO','RULE_TYPE', 'ACTION_ID'])\
                                    .merge(this_dm_name, 
                                           how ='left' , 
                                           on = ['PROM_NO','SUB_DM_ID'])
    
    del this_prom_item,this_prom_cut,this_extra_info,this_dm_name
    
    this_prom_data = this_prom_data.drop_duplicates(["PROM_NO","ITEM_NO"])
    this_prom_data['DATA_SOURCE'] = 'RTMM'
    this_prom_data['PROM_TYPE'] = np.select(
        [this_prom_data['SUB_CHANNEL_ID']==1,
         this_prom_data['SUB_CHANNEL_ID']==2,
         this_prom_data['SUB_CHANNEL_ID']==3
         ],
        ['nDM','nFlyer','HB']
        ) 
    
    this_prom_data['PAGE'] = this_prom_data['PAGE'].fillna('00').map(lambda x : t.fill_zero(int(float(x)),2))
    this_prom_data = this_prom_data.merge(all_item_data)
    
    #去除9處特殊
    this_prom_data = this_prom_data[this_prom_data['DIVISION_ID']!=9]
    
    #增加NEW_CUT_ID方便後續計算
    this_prom_data['NEW_CUT_ID'] = this_prom_data['PAGE'].apply(str)+'_'+this_prom_data['CUT_ID'].apply(str)
    
    
    #對比檔期商品 =================
    print(now()+' '+last_prom_no+' item start')
    sql = ( """
            select distinct promotion_id prom_no, 
            prom_type item_prom_type, channel_id, sub_channel_id,
            item_id item_no, rule_type, action_id
            from pm_st_item@rtmm
            where promotion_id = {}
            and prom_type in ('G') 
            and channel_id in ({})
            and sub_channel_id in ({})
            and store_id in ({}) 
            """).format(last_prom_no,
                        channel,
                        sub_channel,
                        t.list_to_string(store_list)).replace('\n',' ')
        
    last_prom_item= t.query_sql(sql)    
    
    #處理對比檔期CUT_ID提品編號    
    sql = ('''
           select distinct promotion_id prom_no,
           prom_type item_prom_type, channel_id, sub_channel_id,
           item_id item_no, rule_type, action_id, cut_id
           from pm_st_item_plan@rtmm
           where promotion_id = {}
           and prom_type = 'G'
           and channel_id in({})
           and sub_channel_id in ({})
           and store_id in ({})
                ''').format(last_prom_no,
                            channel,
                            sub_channel,
                            t.list_to_string(store_list)).replace('\n',' ') 
    
    last_prom_cut = t.query_sql(sql)
    last_prom_cut['CUT_ID'] = last_prom_cut['CUT_ID'].fillna(0).apply(float).apply(int) #20210203 會議改為0
    
    #處理對比檔期PAGE頁面資訊
    sql = '''
          select promotion_id prom_no, action_id,
          rule_type, sub_dm_id, page
          from pm_st_item_plan_extra@rtmm
          where promotion_id = {}
          '''.format(last_prom_no)
    
    last_extra_info = t.query_sql(sql)
    last_extra_info['PAGE'] = last_extra_info['PAGE'].fillna(0).apply(int) #20210203 會議改為0
            
    #處理對比檔期dm_name_data資訊
    sql = '''
             select promotion_id prom_no,sub_dm_id, sub_dm_name
             from pm_sub_dm_tl@rtmm
             where promotion_id = {}
          '''.format(last_prom_no)
            
    last_dm_name = t.query_sql(sql)
            
    #整合對比檔期促銷商品資訊
    last_prom_data = last_prom_item.merge(last_prom_cut,
                                           how ='left' ,
                                           on = ['PROM_NO', 'ITEM_PROM_TYPE',
                                                 'ACTION_ID','CHANNEL_ID', 'SUB_CHANNEL_ID',
                                                 'ITEM_NO','RULE_TYPE'])\
                                    .merge(last_extra_info,
                                           how ='left' , 
                                           on = ['PROM_NO','RULE_TYPE', 'ACTION_ID'])\
                                    .merge(last_dm_name, 
                                           how ='left' , 
                                           on = ['PROM_NO','SUB_DM_ID'])
    
    del last_prom_item,last_prom_cut,last_extra_info,last_dm_name    
    
    last_prom_data = last_prom_data.drop_duplicates(["PROM_NO","ITEM_NO"])

    last_prom_data['DATA_SOURCE'] = 'RTMM'
    last_prom_data['PROM_TYPE'] = np.select(
        [last_prom_data['SUB_CHANNEL_ID']==1,
         last_prom_data['SUB_CHANNEL_ID']==2,
         last_prom_data['SUB_CHANNEL_ID']==3
         ],
        ['nDM','nFlyer','HB']
        )   
    last_prom_data['PAGE'] = last_prom_data['PAGE'].fillna(0).map(lambda x : t.fill_zero(int(float(x)),2)) #20210203 會議改為0
    last_prom_data = last_prom_data.merge(all_item_data)
    last_prom_data = last_prom_data[last_prom_data['DIVISION_ID']!=9]    
    
    
    # %% #取得檔期銷售資料
    
    ##########################################################################
    #本檔期銷售資料
    ##########################################################################
    print(this_dm_prom_info)
    
    this_start_time = arrow.get(takeout_first(this_dm_prom_info['BEGIN_DATE']))
    this_end_time = arrow.get(takeout_first(this_dm_prom_info['END_DATE']))
    
    period = this_start_time - this_end_time
    period = period.days - 1
    
    
    # 計算前14天的時間點(彈性用)
    # 20210203 會議改為14天；20210305，14天不適用
    this_start_time_before_1 = this_start_time.shift(days=-1)
    this_start_time_before_14 = this_start_time.shift(days=period) 
    
    this_prom_days = (this_end_time-this_start_time).days + 1
    
    # 計算本檔期銷售資料
    print(now(),prom_no,' sales data start.')
    
    
    
    prom_sales = cbyz.get_tnv_qty(begin_date=begin_date,
                                  end_date=end_date,
                                  divisions=[], sections=[],
                                  groups=[], subgroups=[], items=[],
                                  output_level=5, status=None,
                                  exclude_stores=exclude_stores,
                                  include_stores=include_stores,
                                  by_time=True, by_store=False, 
                                  by_member=True,
                                  remove_non_member=False, 
                                  remove_div_9=False,
                                  time_unit='d', time_rename=True,
                                  week_year=True,
                                  remove_uber_eats=False,
                                  database=2,
                                  export_file=True, load_file=False, 
                                  path=path_temp, 
                                  file_name='prom_sales')



    # prom_tickets_overview = cbyz.get_tickets(begin_date=begin_date,
    #                                 end_date=end_date,
    #                                 divisions=[], sections=[], groups=[], 
    #                                 subgroups=[], items=[], output_level=0, 
    #                                 exclude_stores=exclude_stores,
    #                                 include_stores=include_stores,
    #                                 by_store=False, by_member=False,
    #                                 remove_non_member=True, remove_div_9=False,
    #                                 postive_amount_only=True,
    #                                 postive_qty_only=True, pos_under_50_only=True,
    #                                 database=2, export_file=True, load_file=True, 
    #                                 path=path_temp, 
    #                                 file_name='prom_tickets_overview')
    
    
    # prom_tickets_category = cbyz.get_tickets(begin_date=begin_date,
    #                                 end_date=end_date,
    #                                 divisions=[], sections=[], groups=[], 
    #                                 subgroups=[], items=[], output_level=0, 
    #                                 exclude_stores=exclude_stores,
    #                                 include_stores=include_stores,
    #                                 by_store=True, by_member=True,
    #                                 remove_non_member=True, remove_div_9=False,
    #                                 postive_amount_only=True,
    #                                 postive_qty_only=True, pos_under_50_only=True,
    #                                 database=2, export_file=True, load_file=True, 
    #                                 path=path_temp, 
    #                                 file_name='prom_tickets_overview')    


    # last_tickets = cbyz.get_tickets(begin_date=last_begin_date,
    #                              end_date=last_end_date,
    #                             divisions=[], sections=[], groups=[], 
    #                             subgroups=[], items=[], output_level=0, 
    #                             exclude_stores=exclude_stores,
    #                             include_stores=include_stores,
    #                             by_store=True, by_member=True,
    #                             remove_non_member=True, remove_div_9=False,
    #                             postive_amount_only=True,
    #                             postive_qty_only=True, pos_under_50_only=True,
    #                             database=2, export_file=True, load_file=True, 
    #                             path=path_temp, file_name='tickets_1')





    
    this_all_sales_data = pd.DataFrame()
    for store in store_list:
        sql = '''
                select psn_loc_id store_no,psn_id client_no, 
                sell_loc_id buy_store_no,
                pos_rgstr_id pos_no,
                shift_no, shpg_txn_id pk_no,
                to_char(txn_rpt_as_dt, 'yyyymmdd') work_date,
                item_id item_no,item_qty qty,LN_SALE_AMT amount
                from pdata.shpg_txn_ln_st{}@dw
                where to_char(txn_rpt_as_dt,'yyyymmdd') between {} and {}
                '''.format(t.fill_zero(store, 2),
                           this_start_time.format('YYYYMMDD'),
                           this_end_time.format('YYYYMMDD'))
        temp = t.query_sql(sql)
        this_all_sales_data = pd.concat([this_all_sales_data,temp])
        print(store,' end')
        
        

        
    this_g_sales_data = this_all_sales_data.merge(this_prom_data['ITEM_NO'])
        
    #計算本檔期前14天銷售資料
    #20210203 從7天改為14天
    print(now(),prom_no,' sales data bf14 start.')
    
    this_all_sales_data_bf14 = pd.DataFrame()
    for store in store_list:
        sql = '''
                select psn_loc_id store_no,psn_id client_no, 
                sell_loc_id buy_store_no,
                pos_rgstr_id pos_no,
                shift_no, shpg_txn_id pk_no,
                to_char(txn_rpt_as_dt, 'yyyymmdd') work_date,
                item_id item_no,item_qty qty,LN_SALE_AMT amount
                from pdata.shpg_txn_ln_st{}@dw
                where to_char(txn_rpt_as_dt,'yyyymmdd') between {} and {}
                '''.format(t.fill_zero(store, 2),
                           this_start_time_before_14.format('YYYYMMDD'),
                           this_start_time_before_1.format('YYYYMMDD'))
        temp = t.query_sql(sql)
        this_all_sales_data_bf14 = pd.concat([this_all_sales_data_bf14,temp])
        print(store,' end')

    this_g_sales_data_bf14 = this_all_sales_data_bf14.merge(this_prom_data['ITEM_NO'])
    
    #===========================================
    # 計算本檔期總表現    
    # 總來店會員數
    total_member = this_all_sales_data[this_all_sales_data['CLIENT_NO']!=999999]
    total_member = total_member[['STORE_NO','CLIENT_NO']].drop_duplicates()
    total_member = len(total_member)
    
    #this_prom促銷商品會員數
    g_member = this_g_sales_data[this_g_sales_data['CLIENT_NO']!=999999]
    g_member = g_member[['STORE_NO','CLIENT_NO']].drop_duplicates()
    g_member = len(g_member)
    
    #this_prom總業績
    total_amount = this_all_sales_data['AMOUNT'].sum()
    
    #this_prom總來客
    total_ticket = this_all_sales_data.copy()
    total_ticket = total_ticket[total_ticket['POS_NO']<=50]
    total_ticket = total_ticket[total_ticket['AMOUNT']>0]
    total_ticket = total_ticket[total_ticket['QTY']>0]  
    total_ticket['TICKET'] = total_ticket['POS_NO'].apply(str)+'_'+\
                             total_ticket['SHIFT_NO'].apply(str)+'_'+\
                             total_ticket['PK_NO'].apply(str)
    total_ticket = total_ticket['TICKET'].drop_duplicates(keep='first').count()
    
    #this_prom促銷商品業績
    g_amount = this_g_sales_data['AMOUNT'].sum()
    
    #this_prom促銷商品來客
    g_ticket = this_g_sales_data.copy()
    g_ticket = g_ticket[g_ticket['POS_NO']<=50]
    g_ticket = g_ticket[g_ticket['AMOUNT']>0]
    g_ticket = g_ticket[g_ticket['QTY']>0]  
    g_ticket['TICKET'] = g_ticket['POS_NO'].apply(str)+'_'+\
                         g_ticket['SHIFT_NO'].apply(str)+'_'+\
                         g_ticket['PK_NO'].apply(str)
    g_ticket = g_ticket['TICKET'].drop_duplicates(keep='first').count()
    
    #this_prom會員業績
    member_amount = this_all_sales_data.copy()
    member_amount = member_amount[member_amount['CLIENT_NO']!=999999]
    member_amount = member_amount['AMOUNT'].sum()
    
    #this_prom會員來客
    member_ticket = this_all_sales_data.copy()
    member_ticket = member_ticket[member_ticket['CLIENT_NO']!=999999]
    member_ticket = member_ticket[member_ticket['POS_NO']<=50]
    member_ticket = member_ticket[member_ticket['AMOUNT']>0]
    member_ticket = member_ticket[member_ticket['QTY']>0]  
    member_ticket['TICKET'] = member_ticket['POS_NO'].apply(str)+'_'+\
                              member_ticket['SHIFT_NO'].apply(str)+'_'+\
                              member_ticket['PK_NO'].apply(str)
    member_ticket = member_ticket['TICKET'].drop_duplicates(keep='first').count()
    
    #this_prom會員促銷業績
    member_g_amount = this_g_sales_data.copy()
    member_g_amount = member_g_amount[member_g_amount['CLIENT_NO']!=999999]    
    member_g_amount = member_g_amount['AMOUNT'].sum()
    
    #this_prom會員促銷來客
    member_g_ticket = this_g_sales_data.copy()
    member_g_ticket = member_g_ticket[member_g_ticket['CLIENT_NO']!=999999]
    member_g_ticket = member_g_ticket[member_g_ticket['POS_NO']<=50]
    member_g_ticket = member_g_ticket[member_g_ticket['AMOUNT']>0]
    member_g_ticket = member_g_ticket[member_g_ticket['QTY']>0]  
    member_g_ticket['TICKET'] = member_g_ticket['POS_NO'].apply(str)+'_'+\
                              member_g_ticket['SHIFT_NO'].apply(str)+'_'+\
                              member_g_ticket['PK_NO'].apply(str)
    member_g_ticket = member_g_ticket['TICKET'].drop_duplicates(keep='first').count()
    
    #整合
    this_performance = pd.DataFrame([[total_member,g_member,total_amount,total_ticket,
                                      g_amount,g_ticket,
                                      member_amount,member_ticket,
                                      member_g_amount,member_g_ticket]],
                                    columns = ['TOTAL_MEMBER','G_MEMBER',
                                               'TOTAL_AMOUNT','TOTAL_TICKET',
                                               'G_AMOUNT','G_TICKET',
                                               'MEMBER_AMOUNT','MEMBER_TICKET',
                                               'MEMBER_G_AMOUNT','MEMBER_G_TICKET'])    
    
    #計算彈性(開檔後所有天數平均銷售數量/開檔前14天平均銷售數量)
    #原本為「開檔頭7天銷量/開檔前7天銷量」，20210203會議改為 「開檔後所有天數平均銷售數量/開檔前14天平均銷售數量」 
    elv_sales_data_af = this_g_sales_data.copy()
    elv_sales_data_af = elv_sales_data_af[['ITEM_NO','QTY']].groupby('ITEM_NO').sum().reset_index()
    elv_sales_data_af['QTY'] = elv_sales_data_af['QTY']/this_prom_days
        
    elv_sales_data_bf14 = this_g_sales_data_bf14.copy()[['ITEM_NO','QTY']]
    elv_sales_data_bf14 = elv_sales_data_bf14.groupby('ITEM_NO').sum().reset_index()
    elv_sales_data_bf14 = elv_sales_data_bf14.rename(columns = {'QTY':'BF14_QTY'})
    elv_sales_data_bf14['BF14_QTY'] = elv_sales_data_bf14['BF14_QTY']/14
    
    elv_sales_data = elv_sales_data_af.merge(elv_sales_data_bf14,how = 'outer').fillna(0)
    elv_sales_data['ELASTICITY'] = (elv_sales_data['QTY']/elv_sales_data['BF14_QTY'])

    del elv_sales_data_af,elv_sales_data_bf14
    ##########################################################################
    #對比檔期銷售資料
    ##########################################################################
    print(last_dm_prom_info)
    
    last_start_time = arrow.get(takeout_first(last_dm_prom_info['BEGIN_DATE']))
    last_end_time = arrow.get(takeout_first(last_dm_prom_info['END_DATE']))

    last_prom_days = (last_end_time-last_start_time).days+1
    
    #計算對比檔期銷售資料
    #對比檔期之銷售資料，使用dw銷售資料
    #注意，下午的時候因為DW正在轉檔，很可能會跑的很慢
    print(now(),last_prom_no,' sales data start.')

    last_all_sales_data = pd.DataFrame()
    for store in store_list:
        sql = '''
                select psn_loc_id store_no,psn_id client_no, 
                sell_loc_id buy_store_no,
                pos_rgstr_id pos_no,
                shift_no, shpg_txn_id pk_no,
                to_char(txn_rpt_as_dt, 'yyyymmdd') work_date,
                item_id item_no,item_qty qty,LN_SALE_AMT amount
                from pdata.shpg_txn_ln_st{}@dw
                where to_char(txn_rpt_as_dt,'yyyymmdd') between {} and {}
                '''.format(t.fill_zero(store, 2),
                           last_start_time.format('YYYYMMDD'),
                           last_end_time.format('YYYYMMDD'))
        temp = t.query_sql(sql)
        last_all_sales_data = pd.concat([last_all_sales_data,temp])
        print(store,' end')
        
    last_g_sales_data = last_all_sales_data.merge(last_prom_data['ITEM_NO'])
        
    #===========================================
    #計算對比檔期總表現    
    #總來店會員數
    total_member = last_all_sales_data[last_all_sales_data['CLIENT_NO']!=999999]
    total_member = total_member[['STORE_NO','CLIENT_NO']].drop_duplicates()
    total_member = len(total_member)
    
    #last_prom促銷商品會員數
    g_member = last_g_sales_data[last_g_sales_data['CLIENT_NO']!=999999]
    g_member = g_member[['STORE_NO','CLIENT_NO']].drop_duplicates()
    g_member = len(g_member)
    
    #last_prom總業績
    total_amount = last_all_sales_data['AMOUNT'].sum()
    
    #last_prom總來客
    total_ticket = last_all_sales_data.copy()
    total_ticket = total_ticket[total_ticket['POS_NO']<=50]
    total_ticket = total_ticket[total_ticket['AMOUNT']>0]
    total_ticket = total_ticket[total_ticket['QTY']>0]  
    total_ticket['TICKET'] = total_ticket['POS_NO'].apply(str)+'_'+\
                             total_ticket['SHIFT_NO'].apply(str)+'_'+\
                             total_ticket['PK_NO'].apply(str)
    total_ticket = total_ticket['TICKET'].drop_duplicates(keep='first').count()
    
    # last_prom促銷商品業績
    g_amount = last_g_sales_data['AMOUNT'].sum()
    
    #last_prom促銷商品來客
    g_ticket = last_g_sales_data.copy()
    g_ticket = g_ticket[g_ticket['POS_NO']<=50]
    g_ticket = g_ticket[g_ticket['AMOUNT']>0]
    g_ticket = g_ticket[g_ticket['QTY']>0]  
    g_ticket['TICKET'] = g_ticket['POS_NO'].apply(str)+'_'+\
                         g_ticket['SHIFT_NO'].apply(str)+'_'+\
                         g_ticket['PK_NO'].apply(str)
    g_ticket = g_ticket['TICKET'].drop_duplicates(keep='first').count()
    
    # last_prom會員業績
    member_amount = last_all_sales_data.copy()
    member_amount = member_amount[member_amount['CLIENT_NO']!=999999]
    member_amount = member_amount['AMOUNT'].sum()
    
    # last_prom會員來客
    member_ticket = last_all_sales_data.copy()
    member_ticket = member_ticket[member_ticket['CLIENT_NO']!=999999]
    member_ticket = member_ticket[member_ticket['POS_NO']<=50]
    member_ticket = member_ticket[member_ticket['AMOUNT']>0]
    member_ticket = member_ticket[member_ticket['QTY']>0]  
    member_ticket['TICKET'] = member_ticket['POS_NO'].apply(str)+'_'+\
                              member_ticket['SHIFT_NO'].apply(str)+'_'+\
                              member_ticket['PK_NO'].apply(str)
    member_ticket = member_ticket['TICKET'].drop_duplicates(keep='first').count()
    
    # last_prom會員促銷業績
    member_g_amount = last_g_sales_data.copy()
    member_g_amount = member_g_amount[member_g_amount['CLIENT_NO']!=999999]    
    member_g_amount = member_g_amount['AMOUNT'].sum()
    
    # last_prom會員促銷來客
    member_g_ticket = last_g_sales_data.copy()
    member_g_ticket = member_g_ticket[member_g_ticket['CLIENT_NO']!=999999]
    member_g_ticket = member_g_ticket[member_g_ticket['POS_NO']<=50]
    member_g_ticket = member_g_ticket[member_g_ticket['AMOUNT']>0]
    member_g_ticket = member_g_ticket[member_g_ticket['QTY']>0]  
    member_g_ticket['TICKET'] = member_g_ticket['POS_NO'].apply(str)+'_'+\
                              member_g_ticket['SHIFT_NO'].apply(str)+'_'+\
                              member_g_ticket['PK_NO'].apply(str)
    member_g_ticket = member_g_ticket['TICKET'].drop_duplicates(keep='first').count()
    
    # 整合
    last_performance = pd.DataFrame([[total_member,g_member,total_amount,total_ticket,
                                      g_amount,g_ticket,
                                      member_amount,member_ticket,
                                      member_g_amount,member_g_ticket]],
                                    columns = ['TOTAL_MEMBER','G_MEMBER',
                                               'TOTAL_AMOUNT','TOTAL_TICKET',
                                               'G_AMOUNT','G_TICKET',
                                               'MEMBER_AMOUNT','MEMBER_TICKET',
                                               'MEMBER_G_AMOUNT','MEMBER_G_TICKET'])      

    #==============================================================================
    #輸出表格
    export_permance_total = {
        '本檔期整體業績': t.format_thou(takeout_first(this_performance['TOTAL_AMOUNT'])),
        #'對比檔期整體業績' : t.format_thou(takeout_first(last_performance['TOTAL_AMOUNT'])),
        '本檔期整體來客': t.format_thou(takeout_first(this_performance['TOTAL_TICKET'])),
        #'對比檔期整體來客': t.format_thou(takeout_first(last_performance['TOTAL_TICKET'])),
        '本檔期不重複會員數' : t.format_thou(takeout_first(this_performance['TOTAL_MEMBER'])),
        #'對比檔期不重複會員數' : t.format_thou(takeout_first(last_performance['TOTAL_MEMBER'])),        
        #'業績成長率(%)' : round((takeout_first(this_performance['TOTAL_AMOUNT'])-takeout_first(last_performance['TOTAL_AMOUNT']))/takeout_first(last_performance['TOTAL_AMOUNT'])*100,2),
        #'來客成長率(%)' : round((takeout_first(this_performance['TOTAL_TICKET'])-takeout_first(last_performance['TOTAL_TICKET']))/takeout_first(last_performance['TOTAL_TICKET'])*100,2),
        #'不重複會員數成長率(%)' : round((takeout_first(this_performance['TOTAL_MEMBER'])-takeout_first(last_performance['TOTAL_MEMBER']))/takeout_first(last_performance['TOTAL_MEMBER'])*100,2),
        '本檔期平均日業績' : t.format_thou(takeout_first(this_performance['TOTAL_AMOUNT'])/this_prom_days),
        '本檔期平均每日來客數' : t.format_thou(takeout_first(this_performance['TOTAL_TICKET'])/this_prom_days),
        '本檔期客單價' : t.format_thou(takeout_first(this_performance['TOTAL_AMOUNT']) / takeout_first(this_performance['TOTAL_TICKET'])),
        #'對比檔期客單價' : t.format_thou(takeout_first(last_performance['TOTAL_AMOUNT']) / takeout_first(last_performance['TOTAL_TICKET'])),
        }
    
    export_permance_total = pd.DataFrame.from_dict(export_permance_total,orient='index')
    export_permance_total['INDEX'] = export_permance_total.index
    export_permance_total = export_permance_total[['INDEX',0]]
    
    export_permance_g = {
        '本檔期轉G商品業績' : t.format_thou(takeout_first(this_performance['G_AMOUNT'])),
        #'對比檔期轉G商品業績' : t.format_thou(takeout_first(last_performance['G_AMOUNT'])),
        '本檔期轉G商品來客' : t.format_thou(takeout_first(this_performance['G_TICKET'])),
        #'對比檔期轉G商品來客' : t.format_thou(takeout_first(last_performance['G_TICKET'])),
        '本檔期轉G商品平均日業績' : t.format_thou(takeout_first(this_performance['G_AMOUNT'])/this_prom_days),
        '本檔期轉G商品平均每日來客數' : t.format_thou(takeout_first(this_performance['G_TICKET'])/this_prom_days),
        '本檔期轉G商品客單價' : t.format_thou(takeout_first(this_performance['G_AMOUNT']) / takeout_first(this_performance['G_TICKET'])),
        '本檔期轉G商品不重複會員數' : t.format_thou(takeout_first(this_performance['G_MEMBER'])),
        'DM佔總體業績(%)' :  round(takeout_first(this_performance['G_AMOUNT'])/takeout_first(this_performance['TOTAL_AMOUNT'])*100,2)
        #'對比檔期轉G商品不重複會員數' : t.format_thou(takeout_first(last_performance['G_MEMBER'])),
        #'轉G商品業績成長率(%)' : round((takeout_first(this_performance['G_AMOUNT'])-takeout_first(last_performance['G_AMOUNT']))/takeout_first(last_performance['G_AMOUNT'])*100,2),
        #'轉G商品來客成長率(%)' : round((takeout_first(this_performance['G_TICKET'])-takeout_first(last_performance['G_TICKET']))/takeout_first(last_performance['G_TICKET'])*100,2),
        #'轉G商品不重複會員數成長率(%)': round((takeout_first(this_performance['G_MEMBER'])-takeout_first(last_performance['G_MEMBER']))/takeout_first(last_performance['G_MEMBER'])*100,2),
        }
    export_permance_g = pd.DataFrame.from_dict(export_permance_g,orient='index')
    export_permance_g['INDEX'] = export_permance_g.index
    export_permance_g = export_permance_g[['INDEX',0]]

#%% 個別計算個別等級之資訊
     
    #使用到的銷售資料：this_all_sales_data, this_g_sales_data, last_all_sales_data, last_g_sales_data
    #使用到的商品資料：all_item_data,this_prom_data,last_prom_data
    
    #TOTAL Division 處別
    print(now(),'Group_performance Division_total start.')
    division_item_data = all_item_data[['DIVISION_ID','DIVISION_NAME','ITEM_NO']] 
    division_item_data =division_item_data[division_item_data['DIVISION_ID']!=9]
    
    ##本檔期
    division_sales_data = this_all_sales_data.copy().merge(division_item_data)
    
    division_amount = division_sales_data[['DIVISION_ID','AMOUNT']].groupby('DIVISION_ID').sum().reset_index()
    
    division_ticket = division_sales_data[['DIVISION_ID','POS_NO','SHIFT_NO','PK_NO','AMOUNT','QTY']]
    #division_ticket = division_ticket[division_ticket['POS_NO']<=50]
    division_ticket = division_ticket[division_ticket['AMOUNT']>0]
    division_ticket = division_ticket[division_ticket['QTY']>0]
    division_ticket['TICKET'] = division_ticket['POS_NO'].apply(str)+'_'+\
                                division_ticket['SHIFT_NO'].apply(str)+'_'+\
                                division_ticket['PK_NO'].apply(str)
    division_ticket = division_ticket[['DIVISION_ID','TICKET']].drop_duplicates(keep='first').groupby('DIVISION_ID').count().reset_index()
    
    ##對比檔期
    division_last_sales_data = last_all_sales_data.copy().merge(division_item_data)
    
    division_last_amount = division_last_sales_data[['DIVISION_ID','AMOUNT']].groupby('DIVISION_ID').sum().reset_index()
    division_last_amount = division_last_amount.rename(columns = {'AMOUNT' : 'LAST_AMOUNT'})
    
    division_last_ticket = division_last_sales_data[['DIVISION_ID','POS_NO','SHIFT_NO','PK_NO','AMOUNT','QTY']]
    #division_last_ticket = division_last_ticket[division_last_ticket['POS_NO']<=50]
    division_last_ticket = division_last_ticket[division_last_ticket['AMOUNT']>0]
    division_last_ticket = division_last_ticket[division_last_ticket['QTY']>0]
    division_last_ticket['TICKET'] = division_last_ticket['POS_NO'].apply(str)+'_'+\
                                division_last_ticket['SHIFT_NO'].apply(str)+'_'+\
                                division_last_ticket['PK_NO'].apply(str)
    division_last_ticket = division_last_ticket[['DIVISION_ID','TICKET']].drop_duplicates(keep='first').groupby('DIVISION_ID').count().reset_index()
    division_last_ticket = division_last_ticket.rename(columns = {'TICKET' : 'LAST_TICKET'})
    
    #整合資訊
    division_item_data = division_item_data[['DIVISION_ID','DIVISION_NAME']].drop_duplicates().sort_values('DIVISION_ID').reset_index(drop = True)
    
    group_performance_division = division_item_data.merge(division_amount,on = 'DIVISION_ID',how = 'left')\
                                                   .merge(division_ticket,on = 'DIVISION_ID',how = 'left')\
                                                   .merge(division_last_amount,on = 'DIVISION_ID',how = 'left')\
                                                   .merge(division_last_ticket,on = 'DIVISION_ID',how = 'left')
    del division_sales_data,division_last_sales_data,division_item_data,division_amount,division_ticket,division_last_amount,division_last_ticket

    #計算
    group_performance_division['本檔期每日業績'] = (group_performance_division['AMOUNT']/this_prom_days).map(lambda x : t.format_thou(x))
    group_performance_division['對比檔期每日業績'] = (group_performance_division['LAST_AMOUNT']/last_prom_days).map(lambda x : t.format_thou(x))
    group_performance_division['業績成長率(%)'] = round((group_performance_division['AMOUNT']-group_performance_division['LAST_AMOUNT'])/group_performance_division['LAST_AMOUNT']*100,2)
    group_performance_division['本檔期每日來客'] = (group_performance_division['TICKET']/this_prom_days).map(lambda x : t.format_thou(x))
    group_performance_division['對比檔期每日來客'] = (group_performance_division['LAST_TICKET']/last_prom_days).map(lambda x : t.format_thou(x))
    group_performance_division['來客數成長率(%)'] = round((group_performance_division['TICKET']-group_performance_division['LAST_TICKET'])/group_performance_division['LAST_TICKET']*100,2)
            
    group_performance_division['AMOUNT'] = group_performance_division['AMOUNT'].map(lambda x : t.format_thou(x))
    group_performance_division['TICKET'] = group_performance_division['TICKET'].map(lambda x : t.format_thou(x))
    group_performance_division['LAST_AMOUNT'] = group_performance_division['LAST_AMOUNT'].map(lambda x : t.format_thou(x))
    group_performance_division['LAST_TICKET'] = group_performance_division['LAST_TICKET'].map(lambda x : t.format_thou(x))

    group_performance_division['DIVISION_ID'] = np.select(
        [group_performance_division['DIVISION_ID'] == 1,
         group_performance_division['DIVISION_ID'] == 2,
         group_performance_division['DIVISION_ID'] == 3,
         group_performance_division['DIVISION_ID'] == 4,
         group_performance_division['DIVISION_ID'] == 5
         ],['FR','FMCG','GG','TI','DO']
        )

    #Total Section 課別
    print(now(),'Group_performance Section_total start.')
    section_item_data = all_item_data[['SECTION_ID','SECTION_NAME','ITEM_NO']].sort_values('SECTION_ID')

    ##本檔期
    section_sales_data = this_all_sales_data.copy().merge(section_item_data)
    
    section_amount = section_sales_data[['SECTION_ID','AMOUNT']].groupby('SECTION_ID').sum().reset_index()
    
    section_ticket = section_sales_data[['SECTION_ID','POS_NO','SHIFT_NO','PK_NO','AMOUNT','QTY']]
    #section_ticket = section_ticket[section_ticket['POS_NO']<=50]
    section_ticket = section_ticket[section_ticket['AMOUNT']>0]
    section_ticket = section_ticket[section_ticket['QTY']>0]
    section_ticket['TICKET'] = section_ticket['POS_NO'].apply(str)+'_'+\
                                section_ticket['SHIFT_NO'].apply(str)+'_'+\
                                section_ticket['PK_NO'].apply(str)
    section_ticket = section_ticket[['SECTION_ID','TICKET']].drop_duplicates(keep='first').groupby('SECTION_ID').count().reset_index()
    
    ##對比檔期
    section_last_sales_data = last_all_sales_data.copy().merge(section_item_data)
    
    section_last_amount = section_last_sales_data[['SECTION_ID','AMOUNT']].groupby('SECTION_ID').sum().reset_index()
    section_last_amount = section_last_amount.rename(columns = {'AMOUNT' : 'LAST_AMOUNT'})
    
    section_last_ticket = section_last_sales_data[['SECTION_ID','POS_NO','SHIFT_NO','PK_NO','AMOUNT','QTY']]
    #section_last_ticket = section_last_ticket[section_last_ticket['POS_NO']<=50]
    section_last_ticket = section_last_ticket[section_last_ticket['AMOUNT']>0]
    section_last_ticket = section_last_ticket[section_last_ticket['QTY']>0]
    section_last_ticket['TICKET'] = section_last_ticket['POS_NO'].apply(str)+'_'+\
                                section_last_ticket['SHIFT_NO'].apply(str)+'_'+\
                                section_last_ticket['PK_NO'].apply(str)
    section_last_ticket = section_last_ticket[['SECTION_ID','TICKET']].drop_duplicates(keep='first').groupby('SECTION_ID').count().reset_index()
    section_last_ticket = section_last_ticket.rename(columns = {'TICKET' : 'LAST_TICKET'})
    
    #整合資訊
    sect_item_data = section_item_data[['SECTION_ID','SECTION_NAME']].drop_duplicates().sort_values('SECTION_ID').reset_index(drop = True)
    
    group_performance_section = sect_item_data.merge(section_amount,on = 'SECTION_ID',how = 'left')\
                                                   .merge(section_ticket,on = 'SECTION_ID',how = 'left')\
                                                   .merge(section_last_amount,on = 'SECTION_ID',how = 'left')\
                                                   .merge(section_last_ticket,on = 'SECTION_ID',how = 'left')
    del section_item_data,section_sales_data,section_ticket,section_last_sales_data,section_last_amount,section_last_ticket
    
    #計算
    group_performance_section['本檔期每日業績'] = (group_performance_section['AMOUNT']/this_prom_days).fillna(0).map(lambda x : t.format_thou(x))
    group_performance_section['對比檔期每日業績'] = (group_performance_section['LAST_AMOUNT']/last_prom_days).fillna(0).map(lambda x : t.format_thou(x))
    group_performance_section['業績成長率(%)'] = round((group_performance_section['AMOUNT']-group_performance_section['LAST_AMOUNT'])/group_performance_section['LAST_AMOUNT']*100,2)
    group_performance_section['本檔期每日來客'] = (group_performance_section['TICKET']/this_prom_days).fillna(0).map(lambda x : t.format_thou(x))
    group_performance_section['對比檔期每日來客'] = (group_performance_section['LAST_TICKET']/last_prom_days).fillna(0).map(lambda x : t.format_thou(x))
    group_performance_section['來客數成長率(%)'] = round((group_performance_section['TICKET']-group_performance_section['LAST_TICKET'])/group_performance_section['LAST_TICKET']*100,2)
    
    group_performance_section['AMOUNT'] = group_performance_section['AMOUNT'].fillna(0).map(lambda x : t.format_thou(x))
    group_performance_section['TICKET'] = group_performance_section['TICKET'].fillna(0).map(lambda x : t.format_thou(x))
    group_performance_section['LAST_AMOUNT'] = group_performance_section['LAST_AMOUNT'].fillna(0).map(lambda x : t.format_thou(x))
    group_performance_section['LAST_TICKET'] = group_performance_section['LAST_TICKET'].fillna(0).map(lambda x : t.format_thou(x))

    ######################################S
    #轉G商品 DISIVISION 表現
    print(now(),'Group_performance Division_G start.')    
    ##本檔期
    division_item_data = this_prom_data[['DIVISION_ID','DIVISION_NAME','ITEM_NO']] 
    division_sales_data = this_g_sales_data.copy().merge(division_item_data)
    
    division_amount = division_sales_data[['DIVISION_ID','AMOUNT']].groupby('DIVISION_ID').sum().reset_index()
    
    division_ticket = division_sales_data[['DIVISION_ID','POS_NO','SHIFT_NO','PK_NO','AMOUNT','QTY']]
    #division_ticket = division_ticket[division_ticket['POS_NO']<=50]
    division_ticket = division_ticket[division_ticket['AMOUNT']>0]
    division_ticket = division_ticket[division_ticket['QTY']>0]
    division_ticket['TICKET'] = division_ticket['POS_NO'].apply(str)+'_'+\
                                division_ticket['SHIFT_NO'].apply(str)+'_'+\
                                division_ticket['PK_NO'].apply(str)
    division_ticket = division_ticket[['DIVISION_ID','TICKET']].drop_duplicates(keep='first').groupby('DIVISION_ID').count().reset_index()
    
    ##對比檔期
    division_item_data = last_prom_data[['DIVISION_ID','DIVISION_NAME','ITEM_NO']] 
    division_last_sales_data = last_g_sales_data.copy().merge(division_item_data)
    
    division_last_amount = division_last_sales_data[['DIVISION_ID','AMOUNT']].groupby('DIVISION_ID').sum().reset_index()
    division_last_amount = division_last_amount.rename(columns = {'AMOUNT' : 'LAST_AMOUNT'})
    
    division_last_ticket = division_last_sales_data[['DIVISION_ID','POS_NO','SHIFT_NO','PK_NO','AMOUNT','QTY']]
    #division_last_ticket = division_last_ticket[division_last_ticket['POS_NO']<=50]
    division_last_ticket = division_last_ticket[division_last_ticket['AMOUNT']>0]
    division_last_ticket = division_last_ticket[division_last_ticket['QTY']>0]
    division_last_ticket['TICKET'] = division_last_ticket['POS_NO'].apply(str)+'_'+\
                                division_last_ticket['SHIFT_NO'].apply(str)+'_'+\
                                division_last_ticket['PK_NO'].apply(str)
    division_last_ticket = division_last_ticket[['DIVISION_ID','TICKET']].drop_duplicates(keep='first').groupby('DIVISION_ID').count().reset_index()
    division_last_ticket = division_last_ticket.rename(columns = {'TICKET' : 'LAST_TICKET'})

    #計算促銷商品支數
    division_item_number = this_prom_data.copy()
    division_item_number = division_item_number[['DIVISION_ID','NEW_CUT_ID']].drop_duplicates().groupby('DIVISION_ID').count().reset_index()
    
    #計算未達門檻支數
    division_item_noreach = this_prom_data.copy()[['DIVISION_ID','NEW_CUT_ID','ITEM_NO']]
    division_noreach_sales = this_all_sales_data.merge(division_item_noreach)
    
    division_noreach_sales_amount = division_noreach_sales[['DIVISION_ID','NEW_CUT_ID','AMOUNT']].groupby(['DIVISION_ID','NEW_CUT_ID']).sum().reset_index().fillna(0)
    division_noreach_sales_qty = division_noreach_sales[['DIVISION_ID','NEW_CUT_ID','QTY']].groupby(['DIVISION_ID','NEW_CUT_ID']).sum().reset_index().fillna(0)
    
    division_noreach_sales = division_noreach_sales_amount.merge(division_noreach_sales_qty)
    division_noreach_sales['NOREACH'] = np.select(
        [(division_noreach_sales['AMOUNT']<=30000) & (division_noreach_sales['QTY']<=600)],
         [1],default=0)

    division_noreach_sales_amount = division_noreach_sales[division_noreach_sales['NOREACH']==1]
    division_noreach_sales_amount = division_noreach_sales_amount[['DIVISION_ID','AMOUNT']].groupby('DIVISION_ID').mean().reset_index()
    division_noreach_sales_amount = division_noreach_sales_amount.rename(columns = {'AMOUNT' : 'NOREACH_AMOUNT'})
    division_noreach_sales_qty = division_noreach_sales[division_noreach_sales['NOREACH']==1]
    division_noreach_sales_qty = division_noreach_sales_qty[['DIVISION_ID','QTY']].groupby('DIVISION_ID').mean().reset_index()
    division_noreach_sales_qty = division_noreach_sales_qty.rename(columns = {'QTY' : 'NOREACH_QTY'})
    division_noreach_sales = division_noreach_sales[['DIVISION_ID','NOREACH']].groupby('DIVISION_ID').sum().reset_index()

    #計算彈性
    division_item_data = this_prom_data[['DIVISION_ID','ITEM_NO']] 
    division_elv = elv_sales_data.copy().merge(division_item_data,on = 'ITEM_NO')
    division_elv = division_elv[['DIVISION_ID','QTY','BF14_QTY']]
    division_elv = division_elv.groupby('DIVISION_ID').sum().reset_index()
    division_elv['ELASTICITY'] = division_elv['QTY']/division_elv['BF14_QTY']
    division_elv = division_elv[['DIVISION_ID','ELASTICITY']]
    
    #整合資訊
    division_item_data = all_item_data[['DIVISION_ID','DIVISION_NAME']].drop_duplicates().sort_values('DIVISION_ID').reset_index(drop = True)
    
    group_performance_division_G = division_item_data.merge(division_amount,on = 'DIVISION_ID',how = 'left')\
                                                   .merge(division_ticket,on = 'DIVISION_ID',how = 'left')\
                                                   .merge(division_last_amount,on = 'DIVISION_ID',how = 'left')\
                                                   .merge(division_last_ticket,on = 'DIVISION_ID',how = 'left')\
                                                   .merge(division_item_number,on = 'DIVISION_ID',how = 'left')\
                                                   .merge(division_noreach_sales,on = 'DIVISION_ID',how = 'left')\
                                                   .merge(division_elv,on = 'DIVISION_ID',how = 'left')\
                                                   .merge(division_noreach_sales_amount,on = 'DIVISION_ID',how = 'left')\
                                                   .merge(division_noreach_sales_qty,on = 'DIVISION_ID',how = 'left')

    del division_noreach_sales_amount,division_noreach_sales_qty,division_elv,division_item_noreach,division_noreach_sales,division_sales_data,division_last_sales_data,division_item_data,division_amount,division_ticket,division_last_amount,division_last_ticket,division_item_number

    #計算
    group_performance_division_G['本檔期每日業績'] = (group_performance_division_G['AMOUNT']/this_prom_days).fillna(0).map(lambda x : t.format_thou(x))
    group_performance_division_G['對比檔期每日業績'] = (group_performance_division_G['LAST_AMOUNT']/last_prom_days).fillna(0).map(lambda x : t.format_thou(x))
    group_performance_division_G['業績成長率(%)'] = round((group_performance_division_G['AMOUNT']-group_performance_division_G['LAST_AMOUNT'])/group_performance_division_G['LAST_AMOUNT']*100,2)
    group_performance_division_G['本檔期每日來客'] = (group_performance_division_G['TICKET']/this_prom_days).fillna(0).map(lambda x : t.format_thou(x))
    group_performance_division_G['對比檔期每日來客'] = (group_performance_division_G['LAST_TICKET']/last_prom_days).fillna(0).map(lambda x : t.format_thou(x))
    group_performance_division_G['來客數成長率(%)'] = round((group_performance_division_G['TICKET']-group_performance_division_G['LAST_TICKET'])/group_performance_division_G['LAST_TICKET']*100,2)

    group_performance_division_G['ELASTICITY'] = group_performance_division_G['ELASTICITY'].fillna(0).map(lambda x : round(x,2))
    group_performance_division_G['NOREACH_AMOUNT'] = group_performance_division_G['NOREACH_AMOUNT'].fillna(0).map(lambda x : t.format_thou(x))
    group_performance_division_G['NOREACH_QTY'] = group_performance_division_G['NOREACH_QTY'].fillna(0).map(lambda x : t.format_thou(x))
    
    group_performance_division_G = group_performance_division_G.rename(columns = {'NEW_CUT_ID' : '支數',
                                                                                  'NOREACH' : '未達門檻支數',
                                                                                  'ELASTICITY' : '彈性',
                                                                                  'NOREACH_AMOUNT' : '未達門檻支數業績',
                                                                                  'NOREACH_QTY' : '未達門檻支數銷量'
                                                                                  })
    
    group_performance_division_G['DIVISION_ID'] = np.select(
        [group_performance_division_G['DIVISION_ID'] == 1,
         group_performance_division_G['DIVISION_ID'] == 2,
         group_performance_division_G['DIVISION_ID'] == 3,
         group_performance_division_G['DIVISION_ID'] == 4,
         group_performance_division_G['DIVISION_ID'] == 5
         ],['FR','FMCG','GG','TI','DO']
        )    
    group_performance_division_G = group_performance_division_G.fillna(0)


    #轉G商品 SECTION 表現
    print(now(),'Group_performance Section_G start.')
    section_item_data = this_prom_data[['SECTION_ID','SECTION_NAME','ITEM_NO']].sort_values('SECTION_ID')

    ##本檔期
    section_sales_data = this_g_sales_data.copy().merge(section_item_data)
    
    section_amount = section_sales_data[['SECTION_ID','AMOUNT']].groupby('SECTION_ID').sum().reset_index()
    
    section_ticket = section_sales_data[['SECTION_ID','POS_NO','SHIFT_NO','PK_NO','AMOUNT','QTY']]
    #section_ticket = section_ticket[section_ticket['POS_NO']<=50]
    section_ticket = section_ticket[section_ticket['AMOUNT']>0]
    section_ticket = section_ticket[section_ticket['QTY']>0]
    section_ticket['TICKET'] = section_ticket['POS_NO'].apply(str)+'_'+\
                                section_ticket['SHIFT_NO'].apply(str)+'_'+\
                                section_ticket['PK_NO'].apply(str)
    section_ticket = section_ticket[['SECTION_ID','TICKET']].drop_duplicates(keep='first').groupby('SECTION_ID').count().reset_index()
    
    ##對比檔期
    section_item_data = last_prom_data[['SECTION_ID','SECTION_NAME','ITEM_NO']].sort_values('SECTION_ID')
    section_last_sales_data = last_g_sales_data.copy().merge(section_item_data)
    
    section_last_amount = section_last_sales_data[['SECTION_ID','AMOUNT']].groupby('SECTION_ID').sum().reset_index()
    section_last_amount = section_last_amount.rename(columns = {'AMOUNT' : 'LAST_AMOUNT'})
    
    section_last_ticket = section_last_sales_data[['SECTION_ID','POS_NO','SHIFT_NO','PK_NO','AMOUNT','QTY']]
    #section_last_ticket = section_last_ticket[section_last_ticket['POS_NO']<=50]
    section_last_ticket = section_last_ticket[section_last_ticket['AMOUNT']>0]
    section_last_ticket = section_last_ticket[section_last_ticket['QTY']>0]
    section_last_ticket['TICKET'] = section_last_ticket['POS_NO'].apply(str)+'_'+\
                                section_last_ticket['SHIFT_NO'].apply(str)+'_'+\
                                section_last_ticket['PK_NO'].apply(str)
    section_last_ticket = section_last_ticket[['SECTION_ID','TICKET']].drop_duplicates(keep='first').groupby('SECTION_ID').count().reset_index()
    section_last_ticket = section_last_ticket.rename(columns = {'TICKET' : 'LAST_TICKET'})
    
    #計算促銷商品支數
    section_item_number = this_prom_data.copy()
    section_item_number = section_item_number[['SECTION_ID','NEW_CUT_ID']].drop_duplicates().groupby('SECTION_ID').count().reset_index()
    
    #計算未達門檻支數
    section_item_noreach = this_prom_data.copy()[['SECTION_ID','NEW_CUT_ID','ITEM_NO']]
    section_item_noreach_sales = this_all_sales_data.merge(section_item_noreach)
    
    section_noreach_sales_amount = section_item_noreach_sales[['SECTION_ID','NEW_CUT_ID','AMOUNT']].groupby(['SECTION_ID','NEW_CUT_ID']).sum().reset_index().fillna(0)
    section_noreach_sales_qty = section_item_noreach_sales[['SECTION_ID','NEW_CUT_ID','QTY']].groupby(['SECTION_ID','NEW_CUT_ID']).sum().reset_index().fillna(0)
    
    section_item_noreach_sales = section_noreach_sales_amount.merge(section_noreach_sales_qty)
    section_item_noreach_sales['NOREACH'] = np.select(
        [(section_item_noreach_sales['AMOUNT']<=30000) & (section_item_noreach_sales['QTY']<=600)],
         [1],default=0)

    section_item_noreach_sales_amount = section_item_noreach_sales[section_item_noreach_sales['NOREACH']==1]
    section_item_noreach_sales_amount = section_item_noreach_sales_amount[['SECTION_ID','AMOUNT']].groupby('SECTION_ID').mean().reset_index()
    section_item_noreach_sales_amount = section_item_noreach_sales_amount.rename(columns = {'AMOUNT' : 'NOREACH_AMOUNT'})
    section_item_noreach_sales_qty = section_item_noreach_sales[section_item_noreach_sales['NOREACH']==1]
    section_item_noreach_sales_qty = section_item_noreach_sales_qty[['SECTION_ID','QTY']].groupby('SECTION_ID').mean().reset_index()
    section_item_noreach_sales_qty = section_item_noreach_sales_qty.rename(columns = {'QTY' : 'NOREACH_QTY'})
    section_item_noreach_sales = section_item_noreach_sales[['SECTION_ID','NOREACH']].groupby('SECTION_ID').sum().reset_index()

    #計算彈性
    section_item_data = this_prom_data[['SECTION_ID','ITEM_NO']] 
    section_elv = elv_sales_data.copy().merge(section_item_data,on = 'ITEM_NO')
    section_elv = section_elv[['SECTION_ID','QTY','BF14_QTY']]
    section_elv = section_elv.groupby('SECTION_ID').sum().reset_index()
    section_elv['ELASTICITY'] = section_elv['QTY']/section_elv['BF14_QTY']
    section_elv = section_elv[['SECTION_ID','ELASTICITY']]    
    
    #整合資訊
    sect_item_data = all_item_data[['SECTION_ID','SECTION_NAME']].drop_duplicates().sort_values('SECTION_ID').reset_index(drop = True)
    
    group_performance_section_G = sect_item_data.merge(section_amount,on = 'SECTION_ID',how = 'left')\
                                                   .merge(section_ticket,on = 'SECTION_ID',how = 'left')\
                                                   .merge(section_last_amount,on = 'SECTION_ID',how = 'left')\
                                                   .merge(section_last_ticket,on = 'SECTION_ID',how = 'left')\
                                                   .merge(section_item_number,on = 'SECTION_ID',how = 'left')\
                                                   .merge(section_item_noreach_sales,on = 'SECTION_ID',how = 'left')\
                                                   .merge(section_elv,on = 'SECTION_ID',how = 'left')\
                                                   .merge(section_item_noreach_sales_amount,on = 'SECTION_ID',how = 'left')\
                                                   .merge(section_item_noreach_sales_qty,on = 'SECTION_ID',how = 'left')
    
    
    del section_item_number,section_item_noreach_sales,section_elv,section_item_noreach_sales_amount,section_item_noreach_sales_qty,section_item_data,section_sales_data,section_ticket,section_last_sales_data,section_last_amount,section_last_ticket
    
    #計算
    group_performance_section_G['本檔期每日業績'] = (group_performance_section_G['AMOUNT']/this_prom_days).fillna(0).map(lambda x : t.format_thou(x))
    group_performance_section_G['對比檔期每日業績'] = (group_performance_section_G['LAST_AMOUNT']/last_prom_days).fillna(0).map(lambda x : t.format_thou(x))
    group_performance_section_G['業績成長率(%)'] = round((group_performance_section_G['AMOUNT']-group_performance_section_G['LAST_AMOUNT'])/group_performance_section_G['LAST_AMOUNT']*100,2)
    group_performance_section_G['本檔期每日來客'] = (group_performance_section_G['TICKET']/this_prom_days).fillna(0).map(lambda x : t.format_thou(x))
    group_performance_section_G['對比檔期每日來客'] = (group_performance_section_G['LAST_TICKET']/last_prom_days).fillna(0).map(lambda x : t.format_thou(x)) 
    group_performance_section_G['來客數成長率(%)'] = round((group_performance_section_G['TICKET']-group_performance_section_G['LAST_TICKET'])/group_performance_section_G['LAST_TICKET']*100,2)
    
    group_performance_section_G['ELASTICITY'] = group_performance_section_G['ELASTICITY'].fillna(0).map(lambda x : round(x,2))
    group_performance_section_G['NOREACH_AMOUNT'] = group_performance_section_G['NOREACH_AMOUNT'].fillna(0).map(lambda x : t.format_thou(x))
    group_performance_section_G['NOREACH_QTY'] = group_performance_section_G['NOREACH_QTY'].fillna(0).map(lambda x : t.format_thou(x))
    
    group_performance_section_G.rename(columns = {'NEW_CUT_ID' : '支數',
                                                   'NOREACH' : '未達門檻支數',
                                                   'ELASTICITY' : '彈性',
                                                   'NOREACH_AMOUNT' : '未達門檻支數業績',
                                                   'NOREACH_QTY' : '未達門檻支數銷量'
                                                   })
    
    
    group_performance_section_G = group_performance_section_G.fillna(0)
     
    #%%page_view
    #使用資料this_prom_data
    print(now(),'page view start')
    page_view_sales_data = this_g_sales_data.copy()
    page_view_sales_data = page_view_sales_data.merge(this_prom_data)
    
    #檔期業績
    page_amount = page_view_sales_data[['PAGE','AMOUNT']].groupby('PAGE').sum().reset_index()
    
    #檔期業績(每日)
    page_amount['AMOUNT_PER_DAY'] = page_amount['AMOUNT']/this_prom_days
    
    #檔期來客數
    page_ticket = page_view_sales_data[['PAGE','POS_NO','SHIFT_NO','PK_NO','AMOUNT','QTY']]
    #page_ticket = page_ticket[page_ticket['POS_NO']<=50]
    page_ticket = page_ticket[page_ticket['AMOUNT']>0]
    page_ticket = page_ticket[page_ticket['QTY']>0]
    page_ticket['TICKET'] = page_ticket['POS_NO'].apply(str)+'_'+\
                            page_ticket['SHIFT_NO'].apply(str)+'+'+\
                            page_ticket['PK_NO'].apply(str)
    page_ticket = page_ticket[['PAGE','TICKET']].drop_duplicates(keep='first').groupby('PAGE').count().reset_index()
    
    #商品支數
    page_item_num = this_prom_data.copy()[['PAGE','NEW_CUT_ID']].drop_duplicates().groupby('PAGE').count().reset_index()
    
    #未達門檻商品支數
    page_noreach = page_view_sales_data[['PAGE','NEW_CUT_ID','AMOUNT','QTY']]
    page_noreach = page_noreach.groupby(['PAGE','NEW_CUT_ID']).sum().reset_index()
    page_noreach = page_noreach[(page_noreach['AMOUNT']<= 30000) & (page_noreach['QTY']<= 600)]
    page_noreach = page_noreach[['PAGE','NEW_CUT_ID']].groupby('PAGE').count().reset_index()
    page_noreach = page_noreach.rename(columns = {'NEW_CUT_ID':'ITEM_CUT_NUM' })
    
    #檔期主題
    page_num = list(this_prom_data['PAGE'].drop_duplicates().sort_values())
    main_theme = this_prom_data[['PAGE','SUB_DM_NAME']].drop_duplicates().fillna('NOSUB').to_records(index=False)
    main_theme_dict = {}
    
    for i in main_theme:
        if i[0] in main_theme_dict.keys():
            main_theme_dict[i[0]]+='/'+i[1]
        else:
            main_theme_dict[i[0]]=i[1]
    
    main_theme = pd.DataFrame.from_dict(main_theme_dict,orient='index').rename(columns={0:'theme'})
    main_theme['PAGE'] = main_theme.index
    main_theme = main_theme.sort_values('PAGE')

    ndm_gproduct_page = this_prom_data.copy()
    ndm_gproduct_page = df_condition('ndm_gproduct_page','PROM_NO','==int(prom_no)')[['PAGE','ITEM_NO']]
    ndm_gproduct_page = ndm_gproduct_page.merge(all_item_data[['SECTION_ID','ITEM_NO']],how = 'left')
    ndm_gproduct_page = ndm_gproduct_page.drop_duplicates()
    ndm_gproduct_page = ndm_gproduct_page.drop_duplicates(['PAGE','SECTION_ID'])
    ndm_gproduct_page['SECTION_ID'] = ndm_gproduct_page['SECTION_ID'].apply(str)
    ndm_gproduct_page = ndm_gproduct_page.sort_values(['SECTION_ID'])
    ndm_gproduct_page = ndm_gproduct_page.groupby(['PAGE'])
    ndm_gproduct_page = ndm_gproduct_page['SECTION_ID'].apply(','.join).reset_index().rename(columns = {'SECTION_ID' : 'HAVE_SECTION' })
     
    #整合資訊
    page_info = main_theme.merge(page_amount,how = 'left')\
                          .merge(page_ticket,how = 'left')\
                          .merge(page_item_num,how = 'left')\
                          .merge(page_noreach,how = 'left')\
                          .merge(ndm_gproduct_page,how = 'left')
    
    page_info = page_info[['PAGE','theme','HAVE_SECTION','AMOUNT','TICKET','AMOUNT_PER_DAY','NEW_CUT_ID','ITEM_CUT_NUM']].fillna(0)
    page_info = page_info.rename(columns = {'PAGE':'頁數',
                                            'HAVE_SECTION' : '所含課別',
                                            'theme' : '檔期主題',
                                            'AMOUNT' : '檔期業績',
                                            'TICKET' : '檔期來客數',
                                            'AMOUNT_PER_DAY' : '檔期業績(每日)',
                                            'NEW_CUT_ID':'支數',
                                            'ITEM_CUT_NUM' : '未達門檻支數'
                                            })
    
    # page_info.to_excel(file_path+'page_info.xlsx',encoding = 'ANSI',index = False) 
    #################################################################
    #page_item
    page_item_item = this_prom_data.copy()
    page_item_item = page_item_item[['PAGE','DIVISION_ID',
                                     'SECTION_ID','SECTION_EN_NAME',
                                     'CUT_ID','CLUSTER_ID','CLUSTER_NAME',
                                     'SERIES_ID','SERIES_NAME',
                                     'ITEM_NO','ITEM_NAME']]
    
    page_item_g_data = this_g_sales_data.copy()
    
    page_item_item_amount = this_g_sales_data[['ITEM_NO','AMOUNT']].groupby('ITEM_NO').sum().reset_index()
    
    page_item_item_qty = this_g_sales_data[['ITEM_NO','QTY']].groupby('ITEM_NO').sum().reset_index()
    
    page_item_item_ticket = this_g_sales_data[['ITEM_NO','POS_NO','SHIFT_NO','PK_NO']]
    #page_item_item_ticket = page_item_item_ticket[page_item_item_ticket['POS_NO']<=50]
    page_item_item_ticket = page_item_item_ticket[page_item_item_ticket['SHIFT_NO']>0]
    page_item_item_ticket = page_item_item_ticket[page_item_item_ticket['PK_NO']>0]
    page_item_item_ticket['TICKET'] = page_item_item_ticket['POS_NO'].apply(str)+'_'+\
                                      page_item_item_ticket['SHIFT_NO'].apply(str)+'_'+\
                                      page_item_item_ticket['PK_NO'].apply(str)
    
    page_item_item_ticket = page_item_item_ticket[['ITEM_NO','TICKET']].drop_duplicates(keep='first').groupby('ITEM_NO').count().reset_index()
    
    page_item_elv = elv_sales_data.copy()
    page_item_elv = page_item_elv[page_item_elv['ITEM_NO'].isin(list(page_item_item['ITEM_NO']))]
    page_item_elv = page_item_elv[['ITEM_NO','ELASTICITY']]
    
    page_item_item_final = page_item_item.merge(page_item_item_amount,how = 'left')\
                                         .merge(page_item_item_qty,how = 'left')\
                                         .merge(page_item_item_ticket,how = 'left')\
                                         .merge(page_item_elv, how = 'left')
    
    page_item_item_final['NO_REACH'] = np.select(
        [(page_item_item_final['AMOUNT']<=30000)&(page_item_item_final['QTY']<=600)],
        ['Y'],'N')        
    
    page_item_item_final['AMOUNT'] = page_item_item_final['AMOUNT'].fillna(0).map(lambda x : t.format_thou(x))
    page_item_item_final['TICKET'] = page_item_item_final['TICKET'].fillna(0).map(lambda x : t.format_thou(x))  
    page_item_item_final['QTY'] = page_item_item_final['QTY'].fillna(0).map(lambda x : t.format_thou(x))
    page_item_item_final['ELASTICITY'] = page_item_item_final['ELASTICITY'].fillna(0).map(lambda x : round(x,2))

    page_item_item_final['DIVISION_ID'] = np.select(
        [page_item_item_final['DIVISION_ID'] == 1,
         page_item_item_final['DIVISION_ID'] == 2,
         page_item_item_final['DIVISION_ID'] == 3,
         page_item_item_final['DIVISION_ID'] == 4,
         page_item_item_final['DIVISION_ID'] == 5
         ],['FR','FMCG','GG','TI','DO']
        )

    # page_item_item_final.to_excel(file_path+'page_item_items.xlsx',encoding = 'utf8',index = False)
    
    #################################################################
    #page_item_cut
    #頁數	課別代號	課別_提品編號	檔期業績	檔期來客數	檔期商品銷量	彈性	未達門檻商品
    page_item_cut = this_prom_data.copy()
    page_item_cut = page_item_cut[['PAGE','SECTION_ID','SECTION_EN_NAME','CUT_ID','ITEM_NO']]
    
    page_item_g_data = this_g_sales_data.copy()
    page_item_g_data = page_item_g_data.merge(page_item_cut[['PAGE','CUT_ID','ITEM_NO']])
    page_item_g_data = page_item_g_data.drop('ITEM_NO',axis = 1)
    
    page_item_item_amount = page_item_g_data[['PAGE','CUT_ID','AMOUNT']].groupby(['PAGE','CUT_ID']).sum().reset_index()
    
    page_item_item_qty = page_item_g_data[['PAGE','CUT_ID','QTY']].groupby(['PAGE','CUT_ID']).sum().reset_index()
    
    page_item_item_ticket = page_item_g_data[['PAGE','CUT_ID','POS_NO','SHIFT_NO','PK_NO']]
    #page_item_item_ticket = page_item_item_ticket[page_item_item_ticket['POS_NO']<=50]
    page_item_item_ticket = page_item_item_ticket[page_item_item_ticket['SHIFT_NO']>0]
    page_item_item_ticket = page_item_item_ticket[page_item_item_ticket['PK_NO']>0]
    page_item_item_ticket['TICKET'] = page_item_item_ticket['POS_NO'].apply(str)+'_'+\
                                      page_item_item_ticket['SHIFT_NO'].apply(str)+'_'+\
                                      page_item_item_ticket['PK_NO'].apply(str)
    
    page_item_item_ticket = page_item_item_ticket[['PAGE','CUT_ID','TICKET']].drop_duplicates(keep='first').groupby(['PAGE','CUT_ID']).count().reset_index()
    
    page_item_elv = elv_sales_data.copy()
    page_item_elv = page_item_elv[page_item_elv['ITEM_NO'].isin(list(page_item_cut['ITEM_NO']))]
    page_item_elv = page_item_elv.merge(page_item_cut)
    page_item_elv = page_item_elv[['PAGE','CUT_ID','QTY','BF14_QTY']]
    page_item_elv = page_item_elv.groupby(['PAGE','CUT_ID']).sum().reset_index()
    page_item_elv['ELASTICITY'] = page_item_elv['QTY']/page_item_elv['BF14_QTY']
    page_item_elv = page_item_elv[['PAGE','CUT_ID','ELASTICITY']]
    
    page_item_cut_final = page_item_cut.drop('ITEM_NO',axis = 1).drop_duplicates()\
        .merge(page_item_item_amount,how = 'left' , on = ['PAGE','CUT_ID'])\
        .merge(page_item_item_qty , how = 'left' , on = ['PAGE','CUT_ID'])\
        .merge(page_item_item_ticket,how = 'left' , on = ['PAGE','CUT_ID'])\
        .merge(page_item_elv,how = 'left' , on = ['PAGE','CUT_ID'])
    
    page_item_cut_final['NO_REACH'] = np.select(
        [(page_item_cut_final['AMOUNT']<=30000)&(page_item_cut_final['QTY']<=600)],
        ['Y'],'N')    
    
    page_item_cut_final['AMOUNT'] = page_item_cut_final['AMOUNT'].fillna(0).map(lambda x : t.format_thou(x))
    page_item_cut_final['QTY'] = page_item_cut_final['QTY'].fillna(0).map(lambda x : t.format_thou(x))    
    page_item_cut_final['TICKET'] = page_item_cut_final['TICKET'].fillna(0).map(lambda x : t.format_thou(x))
    page_item_cut_final['ELASTICITY'] = page_item_cut_final['ELASTICITY'].fillna(0).map(lambda x : round(x,2))
    
    # page_item_cut_final.to_excel(file_path+'page_item_cuts.xlsx',encoding = 'utf8',index = False)
    
    ##########################################################################
    #page_item_series
    page_item_series = this_prom_data.copy()
    page_item_series = page_item_series[['PAGE','DIVISION_ID',
                                         'SECTION_ID','SECTION_EN_NAME',
                                         'RULE_TYPE',
                                         'ITEM_NO','ITEM_NAME',
                                         'SERIES_ID','SERIES_NAME',
                                         'CLUSTER_ID','CLUSTER_NAME'
                                         ]]
        
    page_item_series['NEW_ACTION_ID'] = np.select(
        [page_item_series['RULE_TYPE'] == 'S',
         page_item_series['RULE_TYPE'] == 'C',
         page_item_series['RULE_TYPE'] == 'I'],
        [page_item_series['SERIES_ID'],
         page_item_series['CLUSTER_ID'],
         page_item_series['ITEM_NO']],default = page_item_series['ITEM_NO'] 
        )
    
    page_item_series['NEW_ACTION_NAME'] = np.select(
        [page_item_series['RULE_TYPE'] == 'S',
         page_item_series['RULE_TYPE'] == 'C',
         page_item_series['RULE_TYPE'] == 'I'],
        [page_item_series['SERIES_NAME'],
         page_item_series['CLUSTER_NAME'],
         page_item_series['ITEM_NAME']],default = page_item_series['ITEM_NAME'] 
        )   

    page_item_series['RULE_TYPE'] = np.select(
        [page_item_series['RULE_TYPE'] == 'S',
         page_item_series['RULE_TYPE'] == 'C',
         page_item_series['RULE_TYPE'] == 'B',
         page_item_series['RULE_TYPE'] == 'I'],
        ['系列','群組','組包','單品'],default = '單品'
        )   
    
    #page_item_series = page_item_series[['ITEM_NO','NEW_ACTION_ID','NEW_ACTION_NAME']]
    
    page_item_series_sales = this_g_sales_data.copy().merge(page_item_series,on = 'ITEM_NO')\
                                                     .merge(elv_sales_data[['ITEM_NO','BF14_QTY']],on = 'ITEM_NO')
    page_item_series_sales = page_item_series_sales.drop('ITEM_NO',axis = 1)
    
    page_item_series_amount = page_item_series_sales[['NEW_ACTION_ID','NEW_ACTION_NAME','AMOUNT']].groupby(['NEW_ACTION_ID','NEW_ACTION_NAME']).sum().reset_index()
    
    page_item_series_qty_t = page_item_series_sales[['NEW_ACTION_ID','NEW_ACTION_NAME','QTY']].groupby(['NEW_ACTION_ID','NEW_ACTION_NAME']).sum().reset_index()
    
    page_item_series_ticket = page_item_series_sales[['NEW_ACTION_ID','NEW_ACTION_NAME','POS_NO','SHIFT_NO','PK_NO']]
    #page_item_series_ticket = page_item_series_ticket[page_item_series_ticket['POS_NO']<=50]
    page_item_series_ticket = page_item_series_ticket[page_item_series_ticket['SHIFT_NO']>0]
    page_item_series_ticket = page_item_series_ticket[page_item_series_ticket['PK_NO']>0]
    page_item_series_ticket['TICKET'] = page_item_series_ticket['POS_NO'].apply(str)+'_'+\
                                      page_item_series_ticket['SHIFT_NO'].apply(str)+'_'+\
                                      page_item_series_ticket['PK_NO'].apply(str)
    
    page_item_series_ticket = page_item_series_ticket[['NEW_ACTION_ID','NEW_ACTION_NAME','TICKET']].drop_duplicates().groupby(['NEW_ACTION_ID','NEW_ACTION_NAME']).count().reset_index()
    
    page_item_series_qty = page_item_series_sales.merge(elv_sales_data[['ITEM_NO','BF14_QTY']]).drop('ITEM_NO',axis = 1)
    page_item_series_qty = page_item_series_qty[['NEW_ACTION_ID','NEW_ACTION_NAME','QTY','BF14_QTY']].groupby(['NEW_ACTION_ID','NEW_ACTION_NAME']).sum().reset_index()
    page_item_series_qty['ELASTICITY'] = page_item_series_qty['QTY']/page_item_series_qty['BF14_QTY']
    page_item_series_qty = page_item_series_qty[['NEW_ACTION_ID','NEW_ACTION_NAME','ELASTICITY']]
    
    page_item_series_final = page_item_series.drop(['ITEM_NO','ITEM_NAME'],axis = 1).drop_duplicates()\
        .merge(page_item_series_amount,how = 'left' , on = ['NEW_ACTION_ID','NEW_ACTION_NAME'])\
        .merge(page_item_series_qty_t,how = 'left' , on = ['NEW_ACTION_ID','NEW_ACTION_NAME'])\
        .merge(page_item_series_ticket,how = 'left' , on = ['NEW_ACTION_ID','NEW_ACTION_NAME'])\
        .merge(page_item_series_qty,how = 'left' , on = ['NEW_ACTION_ID','NEW_ACTION_NAME'])
    
    page_item_series_final['NO_REACH'] = np.select(
        [(page_item_series_final['AMOUNT']<=30000)&(page_item_series_final['QTY']<=600)],
        ['Y'],'N')    

    page_item_series_final = page_item_series_final.drop(['SERIES_ID','SERIES_NAME','CLUSTER_ID','CLUSTER_NAME'],axis = 1)
    page_item_series_final = page_item_series_final[~page_item_series_final['NEW_ACTION_ID'].isnull()]

    page_item_series_final['AMOUNT'] = page_item_series_final['AMOUNT'].fillna(0).map(lambda x : t.format_thou(x))
    page_item_series_final['QTY'] = page_item_series_final['QTY'].fillna(0).map(lambda x : t.format_thou(x))    
    page_item_series_final['TICKET'] = page_item_series_final['TICKET'].fillna(0).map(lambda x : t.format_thou(x))
    page_item_series_final['ELASTICITY'] = page_item_series_final['ELASTICITY'].fillna(0).map(lambda x : round(x,2))
    
    #%%TOP10商品
    print(now(),'top 10 item start.')
    top10_data = page_item_series_final.copy()
    top10_data = top10_data[~top10_data['NEW_ACTION_ID'].isnull()].fillna(0)
        
    top10_data_top = top10_data[['DIVISION_ID',
                                 'NEW_ACTION_ID','NEW_ACTION_NAME',
                                 'AMOUNT','TICKET','QTY','ELASTICITY'
                                 ]].sort_values('AMOUNT',ascending = False)
    
    
    top10_data_top_final = pd.DataFrame()
    for i in range(1,6):
        temp = top10_data_top[top10_data_top['DIVISION_ID']==i].head(10)
        top10_data_top_final = pd.concat([top10_data_top_final,temp])
    
    top10_data_low_final = pd.DataFrame()
    for i in range(1,6):
        temp = top10_data_top[top10_data_top['DIVISION_ID']==i].tail(10)
        top10_data_low_final = pd.concat([top10_data_low_final,temp])
    
    #%%計算回應率
    #抓取有寄送DM之顧客紀錄
    print(now(),'client send response rate start.')
    send_dm_sc = t.query_sql(('''
                              select store_no,client_no 
                              from rdba.prom_slip_dm_list
                              where prom_no in ('{}')
                              and condition like '1%'
                                 ''').format(prom_no))
    send_dm_sc['SEND_DM'] = 1
    
    #有收DM且有來店者
    buy_dm_sc = this_all_sales_data[['STORE_NO','CLIENT_NO']].drop_duplicates()
    buy_dm_sc = send_dm_sc.merge(buy_dm_sc,on = ['STORE_NO','CLIENT_NO'])

    #有收DM且有買促銷商品者
    buy_g_sc = this_g_sales_data[['STORE_NO','CLIENT_NO']].drop_duplicates()
    buy_g_sc = send_dm_sc.merge(buy_g_sc,on = ['STORE_NO','CLIENT_NO'])

    #計算
    temp_amount = this_all_sales_data.groupby(['STORE_NO','CLIENT_NO'])['AMOUNT'].sum().reset_index()
    temp_send_dm_amount = send_dm_sc.merge(temp_amount)
    total_temp_send_dm_amount = temp_send_dm_amount['AMOUNT'].sum()
    
    temp_send_g_amount = buy_g_sc.merge(temp_amount)
    total_temp_send_g_amount = temp_send_g_amount['AMOUNT'].sum()
    
    temp_send_dm_response_rate = len(buy_dm_sc)/len(send_dm_sc)*100
    temp_send_g_response_rate = len(buy_g_sc)/len(send_dm_sc)*100
    
    dm_data = pd.DataFrame([[prom_no,
                             total_temp_send_dm_amount,
                             total_temp_send_g_amount,
                             temp_send_dm_response_rate,
                             temp_send_g_response_rate]],
                           columns=['PROM_NO',
                                    'SEND_DM_AMOUNT',
                                    'SEND_G_AMOUNT',
                                    'SEND_DM_RESPONSE_RATE(%)',
                                    'SEND_G_RESPONSE_RATE(%)'])
    
    del temp_amount,temp_send_dm_amount,total_temp_send_dm_amount,temp_send_g_amount,total_temp_send_g_amount
    del temp_send_dm_response_rate,temp_send_g_response_rate

    # dm_data.to_excel(file_path+"dm_data.xlsx")    


    #%% (性別、狀態、星等、年齡區間)
    #顧客檔期前一天狀態
    print(now(),'client send data start.')
    status_flie_name = str(this_start_time_before_1.format('YYYYMMDD'))+'.rds'
    
    
    
    
    if prom_no == '2058':
        status_flie_name = '20201209.rds'
    elif prom_no == '2059':
        status_flie_name = '20201216.rds'
    elif prom_no == '2060':
        status_flie_name = '20201230.rds'
        
    print('status_flie_name為測試版，待更新')
                
        
    member_data = t.readRDS('D:/Data_Mining/Data_Backup/Member_Status/'+status_flie_name)
    
       
    # 顧客的星等狀態 該表每月5日更新
    sql = ('''
        select distinct store_no,client_no,star
        from RDBA.RFM_CUSTOMERVALUE_RANK
        where YEAR_RANK = '{}'
        and MONTH_RANK = '{}'
    ''').replace('\n',' ').format(this_start_time.year,this_start_time.month)
    
    rfm = t.query_sql(sql)
    
    member_data = member_data.merge(rfm, how='left').merge(send_dm_sc)
    del rfm
     
    #計算全業績
    cal_sale_data = this_all_sales_data.copy() \
                    .merge(member_data, on=['STORE_NO','CLIENT_NO'])
                    
    cal_g_sale_data = this_g_sales_data.copy() \
                        .merge(member_data,on = ['STORE_NO','CLIENT_NO'])
    
    cal_title = ['SEX','AGE_LEVEL','STATUS','STAR']

    member_data_integrate = pd.DataFrame()

    for title in cal_title:
        
        
        print(now(),'client send data-',title,' start.')
        send_temp_num = member_data[[title,'SEND_DM']].groupby(title).count().reset_index()
        member_temp_num =  cal_sale_data[['STORE_NO','CLIENT_NO',title,'SEND_DM']].drop_duplicates().groupby([title]).count().reset_index()
        member_temp_num = member_temp_num[[title,'SEND_DM']].rename(columns = {'SEND_DM':'RESPONE'})
        
        member_all_temp_amount = cal_sale_data[[title,'AMOUNT']].groupby([title]).sum().reset_index()
        
        member_all_temp_ticket = cal_sale_data[[title,'POS_NO','SHIFT_NO','PK_NO','AMOUNT','QTY']]
        #member_all_temp_ticket = member_all_temp_ticket[member_all_temp_ticket['POS_NO']<=50]
        member_all_temp_ticket = member_all_temp_ticket[member_all_temp_ticket['AMOUNT']>0]
        member_all_temp_ticket = member_all_temp_ticket[member_all_temp_ticket['QTY']>0]
        member_all_temp_ticket['TICKET'] = member_all_temp_ticket['POS_NO'].apply(str)+'_'+\
                                      member_all_temp_ticket['SHIFT_NO'].apply(str)+'_'+\
                                      member_all_temp_ticket['PK_NO'].apply(str)
        member_all_temp_ticket = member_all_temp_ticket[[title,'TICKET']].drop_duplicates().groupby([title]).count().reset_index()
    
        member_g_temp_amount = cal_g_sale_data[[title,'AMOUNT']].groupby([title]).sum().reset_index()
        member_g_temp_amount = member_g_temp_amount.rename(columns = {'AMOUNT' : 'G_AMOUNT'})
        
        member_g_temp_ticket = cal_g_sale_data[[title,'POS_NO','SHIFT_NO','PK_NO','AMOUNT','QTY']]
        #member_g_temp_ticket = member_g_temp_ticket[member_g_temp_ticket['POS_NO']<=50]
        member_g_temp_ticket = member_g_temp_ticket[member_g_temp_ticket['AMOUNT']>0]
        member_g_temp_ticket = member_g_temp_ticket[member_g_temp_ticket['QTY']>0]
        member_g_temp_ticket['TICKET'] = member_g_temp_ticket['POS_NO'].apply(str)+'_'+\
                                      member_g_temp_ticket['SHIFT_NO'].apply(str)+'_'+\
                                      member_g_temp_ticket['PK_NO'].apply(str)
        member_g_temp_ticket = member_g_temp_ticket[[title,'TICKET']].drop_duplicates().groupby([title]).count().reset_index()
        member_g_temp_ticket = member_g_temp_ticket.rename(columns = {'TICKET' : 'G_TICKET'})
    
        sendlist_temp =  send_temp_num.merge(member_temp_num)\
                                    .merge(member_all_temp_amount)\
                                    .merge(member_all_temp_ticket)\
                                    .merge(member_g_temp_amount)\
                                    .merge(member_g_temp_ticket)
        
        sendlist_temp = sendlist_temp.rename(columns = {title: 'TARGET'})
        sendlist_temp['LABEL'] = title
        
        if title == 'SEX':
            sendlist_temp['TARGET'] = np.select(
                [sendlist_temp['TARGET']==1],
                ['男'],default = '女')
        
        member_data_integrate = pd.concat([member_data_integrate, sendlist_temp])

    # member_data_integrate.to_excel(file_path+'member_send_info.xlsx',index = False)
    
#%%寫入EXCEL   
    wb = Workbook()
    dest_filename = file_path + 'DM_Debrief_'+prom_no+'.xlsx'
    
    ws1 = wb.active
    ws1.title = "overview"
    
    export_prom_info = pd.concat([this_dm_prom_info,last_dm_prom_info])
    export_prom_info = export_prom_info[['PROM_NO','BEGIN_DATE','END_DATE']]
                                  
    export_prom_info['DAYS'] = (export_prom_info['END_DATE'].map(lambda x : arrow.get(x))-\
                                export_prom_info['BEGIN_DATE'].map(lambda x : arrow.get(x))).map(lambda x : x.days)+1
    
    export_prom_info = export_prom_info.rename(columns = {'PROM_NO' : '檔期代號',
                                                          'BEGIN_DATE' : '計算開始時間',
                                                          'END_DATE' : '計算結束時間',
                                                          'DAYS' : '檔期天數'})
        
        
    for r in dataframe_to_rows(export_prom_info, index=False, header=True):
        ws1.append(r)
        
    ws1['A4'] = ' '
    ws1['A5'] = '本檔期期間概況'
    
    for r in dataframe_to_rows(export_permance_total, index=False, header=False):
        ws1.append(r)    
    
    ws1['A12'] = ' '
    ws1['A13'] = '本檔期DM概況(轉G商品)'
    
    for r in dataframe_to_rows(export_permance_g, index=False, header=False):
        ws1.append(r)        
    
    ws1['A22'] = 'DM回應率(%)'
    ws1['B22'] = round(takeout_first(dm_data['SEND_DM_RESPONSE_RATE(%)']),2)
    ws1['A23'] = 'DM商品回應率(%)'
    ws1['B23'] = round(takeout_first(dm_data['SEND_G_RESPONSE_RATE(%)']),2)
    
    ws1['A24'] = ' '
    ws1['A25'] = 'Division overview'
        
    group_performance_division = group_performance_division.rename(columns = 
                                                                   {'DIVISION_ID':'處別代號',
                                                                    'DIVISION_NAME':'處別名稱',
                                                                    'AMOUNT': '本檔期業績',
                                                                    'TICKET' : '本檔期來客',
                                                                    'LAST_AMOUNT' : '對比檔期業績',
                                                                    'LAST_TICKET' : '對比檔期來客'
                                                                    })
    
    group_performance_division = group_performance_division.fillna(0)
    
    for r in dataframe_to_rows(group_performance_division, index=False, header=True):
        ws1.append(r)    
    
    ws1['A32'] = ' '
    ws1['A33'] = 'Section overview'    
    
    
    group_performance_section = group_performance_section.rename(columns = 
                                                                   {'SECTION_ID':'課別代號',
                                                                    'SECTION_NAME':'課別名稱',
                                                                    'AMOUNT': '本檔期業績',
                                                                    'TICKET' : '本檔期來客',
                                                                    'LAST_AMOUNT' : '對比檔期業績',
                                                                    'LAST_TICKET' : '對比檔期來客'
                                                                    })
    
    group_performance_section = group_performance_section.fillna(0)

    for r in dataframe_to_rows(group_performance_section, index=False, header=True):
        ws1.append(r)       
    
    #dm_view
    ws2 = wb.create_sheet(title="DM_view")
    ws2['A1'] = 'Division DM view'
    
    #group_performance_division_G['AMOUNT'] = group_performance_division_G['AMOUNT'].map(lambda x : t.format_thou(x))
    #group_performance_division_G['TICKET'] = group_performance_division_G['TICKET'].map(lambda x : t.format_thou(x))
    #group_performance_division_G['LAST_AMOUNT'] = group_performance_division_G['LAST_AMOUNT'].map(lambda x : t.format_thou(x))
    #group_performance_division_G['LAST_TICKET'] = group_performance_division_G['LAST_TICKET'].map(lambda x : t.format_thou(x))
     
    group_performance_division_G = group_performance_division_G.rename(columns = 
                                                                   {'DIVISION_ID':'處別代號',
                                                                    'DIVISION_NAME':'處別名稱',
                                                                    'AMOUNT': '本檔期業績',
                                                                    'TICKET' : '本檔期來客',
                                                                    'LAST_AMOUNT' : '對比檔期業績',
                                                                    'LAST_TICKET' : '對比檔期來客'
                                                                    })
    
    group_performance_division_G = group_performance_division_G.fillna(0)
    
    for r in dataframe_to_rows(group_performance_division_G, index=False, header=True):
        ws2.append(r)    
    
    ws2['A8'] = ' '
    ws2['A9'] = 'Section DM view'    
    
    #group_performance_section_G['AMOUNT'] = group_performance_section_G['AMOUNT'].map(lambda x : t.format_thou(x))
    #group_performance_section_G['TICKET'] = group_performance_section_G['TICKET'].map(lambda x : t.format_thou(x))
    #group_performance_section_G['LAST_AMOUNT'] = group_performance_section_G['LAST_AMOUNT'].map(lambda x : t.format_thou(x))
    #group_performance_section_G['LAST_TICKET'] = group_performance_section_G['LAST_TICKET'].map(lambda x : t.format_thou(x))
  
    
    group_performance_section_G = group_performance_section_G.rename(columns = 
                                                                   {'SECTION_ID':'課別代號',
                                                                    'SECTION_NAME':'課別名稱',
                                                                    'AMOUNT': '本檔期業績',
                                                                    'TICKET' : '本檔期來客',
                                                                    'LAST_AMOUNT' : '對比檔期業績',
                                                                    'LAST_TICKET' : '對比檔期來客',
                                                                    'NEW_CUT_ID': '支數'	,
                                                                    'NOREACH' :'未達門檻支數',
                                                                    'ELASTICITY' : '彈性',
                                                                    'NOREACH_AMOUNT' : '未達門檻支數業績',
                                                                    'NOREACH_QTY' : '未達門檻支數銷量'
                                                                    })
    
    group_performance_section_G = group_performance_section_G.fillna(0)
    
    for r in dataframe_to_rows(group_performance_section_G, index=False, header=True):
        ws2.append(r)         

    #page_view
    ws3 = wb.create_sheet(title="page_view")
    ws3['A1'] = 'page view'
    
    #page_info['檔期業績'] = page_info['檔期業績'].map(lambda x : t.format_thou(x))
    #page_info['檔期來客數'] = page_info['檔期來客數'].map(lambda x : t.format_thou(x))
    #page_info['檔期業績(每日)'] = page_info['檔期業績(每日)'].map(lambda x : t.format_thou(x))
    
    for r in dataframe_to_rows(page_info, index=False, header=True):
        ws3.append(r)   

    #page_view_item
    ws4 = wb.create_sheet(title="page_item_items")
    ws4['A1'] = 'page_item view'
    
    page_item_item_final = page_item_item_final.rename(columns = { 'PAGE' : '頁數',
                                                                   'DIVISION_ID' : '處別',
                                                                   'SECTION_ID' : '課別',
                                                                   'SECTION_EN_NAME' : '課別英文名',
                                                                   'CUT_ID' : '提品編號',
                                                                   'CLUSTER_ID' : '群組編號',
                                                                   'CLUSTER_NAME' : '群組名稱',
                                                                   'SERIES_ID' : '系列貨號',
                                                                   'SERIES_NAME' : '系列名稱',
                                                                   'ITEM_NO' : '商品貨號',
                                                                   'ITEM_NAME' : '商品名稱',
                                                                   'AMOUNT' : '業績',
                                                                   'TICKET' : '來客數',
                                                                   'QTY' : '銷量',
                                                                   'ELASTICITY' : '彈性',
                                                                   'NO_REACH' : '未達門檻商品'
                                                                   })
    
    for r in dataframe_to_rows(page_item_item_final, index=False, header=True):
        ws4.append(r) 
        
    #page_view_cut
    ws5 = wb.create_sheet(title="page_item_cut")
    ws5['A1'] = 'page_item_cut view'
    
    #page_item_cut_final['AMOUNT'] = page_item_cut_final['AMOUNT'].map(lambda x : t.format_thou(x))
    #page_item_cut_final['TICKET'] = page_item_cut_final['TICKET'].map(lambda x : t.format_thou(x))
    #page_item_cut_final['QTY'] = page_item_cut_final['QTY'].map(lambda x : t.format_thou(x))  
    #page_item_cut_final['ELASTICITY'] = page_item_cut_final['ELASTICITY'].map(lambda x : round(x,2))     
    
    page_item_cut_final = page_item_cut_final.rename(columns = {'PAGE' : '頁數', 
                                                                'SECTION_ID' : '課別', 
                                                                'SECTION_EN_NAME' : '課別英文名', 
                                                                'CUT_ID' : '提品編號', 
                                                                'AMOUNT' : '業績', 
                                                                'QTY' : '銷量',
                                                                'TICKET' : '來客數',
                                                                'ELASTICITY' : '彈性', 
                                                                'NO_REACH' : '未達門檻商品'
        })
    
    for r in dataframe_to_rows(page_item_cut_final, index=False, header=True):
        ws5.append(r) 
        
    #page_view_series
    ws6 = wb.create_sheet(title="page_item_series")
    ws6['A1'] = 'page_item_series view'
    
    #page_item_series_final['AMOUNT'] = page_item_series_final['AMOUNT'].map(lambda x : t.format_thou(x))
    #page_item_series_final['TICKET'] = page_item_series_final['TICKET'].map(lambda x : t.format_thou(x))
    #page_item_series_final['QTY'] = page_item_series_final['QTY'].map(lambda x : t.format_thou(x))  
    #page_item_series_final['ELASTICITY'] = page_item_series_final['ELASTICITY'].map(lambda x : round(x,2))    
    
    page_item_series_final = page_item_series_final.rename(columns = {'PAGE' : '頁數', 
                                                                      'DIVISION_ID' : '處別', 
                                                                      'SECTION_ID' : '課別', 
                                                                      'SECTION_EN_NAME' : '課別英文名', 
                                                                      'RULE_TYPE' : '商品種類',            
                                                                      'NEW_ACTION_ID' : '系列/群組/商品貨號', 
                                                                      'NEW_ACTION_NAME' : '系列/群組/商品名稱', 
                                                                      'AMOUNT' : '業績', 
                                                                      'QTY' : '銷量', 
                                                                      'TICKET' : '來客數',
                                                                      'ELASTICITY' : '彈性', 
                                                                      'NO_REACH' : '未達門檻商品'
                                                                      })
    
    for r in dataframe_to_rows(page_item_series_final, index=False, header=True):
        ws6.append(r)         
        
    #item_view
    ws7 = wb.create_sheet(title="item_view")
    ws7['A1'] = 'Top 10 items by each divion'
    
    #top10_data_top_final['AMOUNT'] = top10_data_top_final['AMOUNT'].map(lambda x : t.format_thou(x))
    #top10_data_top_final['TICKET'] = top10_data_top_final['TICKET'].map(lambda x : t.format_thou(x))
    #top10_data_top_final['QTY'] = top10_data_top_final['QTY'].map(lambda x : t.format_thou(x))  
    #top10_data_top_final['ELASTICITY'] = top10_data_top_final['ELASTICITY'].map(lambda x : round(x,2))       
    
    top10_data_top_final = top10_data_top_final.rename(columns = {'DIVISION_ID' : '處別', 
                                                                  'NEW_ACTION_ID' : '系列/群組/商品貨號', 
                                                                  'NEW_ACTION_NAME' : '系列/群組/商品名稱', 
                                                                  'AMOUNT' : '業績', 
                                                                  'TICKET' : '來客數',       
                                                                  'QTY' : '銷量', 
                                                                  'ELASTICITY' : '彈性'
                                                                  })
    
    for r in dataframe_to_rows(top10_data_top_final, index=False, header=True):
        ws7.append(r)   

    ws7['A53'] = ' '
    ws7['A54'] = 'Low 10 items by each division'

    #top10_data_low_final['AMOUNT'] = top10_data_low_final['AMOUNT'].map(lambda x : t.format_thou(x))
    #top10_data_low_final['TICKET'] = top10_data_low_final['TICKET'].map(lambda x : t.format_thou(x))
    #top10_data_low_final['QTY'] = top10_data_low_final['QTY'].map(lambda x : t.format_thou(x))  
    #top10_data_low_final['ELASTICITY'] = top10_data_low_final['ELASTICITY'].map(lambda x : round(x,2))       
    

    top10_data_low_final = top10_data_low_final.rename(columns = {'DIVISION_ID' : '處別', 
                                                                  'NEW_ACTION_ID' : '系列/群組/商品貨號', 
                                                                  'NEW_ACTION_NAME' : '系列/群組/商品名稱', 
                                                                  'AMOUNT' : '業績', 
                                                                  'TICKET' : '來客數',       
                                                                  'QTY' : '銷量', 
                                                                  'ELASTICITY' : '彈性'
                                                                  })
        
    for r in dataframe_to_rows(top10_data_low_final, index=False, header=True):
        ws7.append(r) 
          
    #send_member
    ws8 = wb.create_sheet(title="send_member")
    ws8['A1'] = '寄送會員表現'
    
    member_data_integrate = member_data_integrate.rename(columns = {'TARGET' : '標籤', 
                                                                  'SEND_DM' : '寄送DM人數', 
                                                                  'RESPONE' : '回應人數', 
                                                                  'AMOUNT' : '業績', 
                                                                  'TICKET' : '來客數',       
                                                                  'QTY' : '銷量', 
                                                                  'ELASTICITY' : '彈性',
                                                                  'G_AMOUNT' : '檔期業績',
                                                                  'G_TICKET' : '檔期來客',
                                                                  'LABEL' : '標籤名稱'
                                                                  })
    
    
    for r in dataframe_to_rows(member_data_integrate, index=False, header=True):
        ws8.append(r)   

    wb.save(filename=dest_filename)

    # %% 後續處理
    # 處理驗證資料上傳
    array=todo_data
    t.get_sheet(url,'驗證用').append_table([array])

    #處理資料寄送
    COMMASPACE	=	',	'

    sender = 'rt.mart.send.mail@gmail.com'
    gmail_password = 'rtmart0000'
    recipients = ['yzwu@rt-mart.com.tw', 
                  'rt.mart.send.mail@gmail.com', 
                  'rt.data.mining01@gmail.com',
                  user_mail
                  ]

    outer = MIMEMultipart()
    outer['Subject'] = 'Debrief'+str(prom_no)
    outer['To'] = COMMASPACE.join(recipients)
    outer['From'] = sender
    outer.preamble = 'You will not see this in a MIME-aware	mail reader.\n'
				
    attachments = [dest_filename]
				
	#加入檔案到MAIL底下
    for file in attachments:
        try:
            with open(file, 'rb') as fp:
                print ('can read faile')
                msg = MIMEBase('application', "octet-stream")
                msg.set_payload(fp.read())
                encoders.encode_base64(msg)
                msg.add_header('Content-Disposition', 'attachment', filename = os.path.basename(file))
                outer.attach(msg)
        except:
            print("Unable to open one of the attachments. Error: ",	sys.exc_info()[0])
            raise
        
        composed = outer.as_string()
        try:
            with smtplib.SMTP('smtp.gmail.com',	587) as s:
                s.ehlo()
                s.starttls()
                s.ehlo()
                s.login(sender,	gmail_password)
                s.sendmail(sender,	recipients,	composed)
                s.close()
                print("Email sent!")
        except:
            print("Unable to send the email. Error: ",	sys.exc_info()[0])
            raise
		    

